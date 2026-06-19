import os, hashlib, secrets, json, smtplib, tempfile, shutil
from datetime import datetime, date, timedelta
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from functools import wraps
from flask import (Flask, render_template_string, request, redirect,
                   url_for, session, flash, jsonify, send_file)
from werkzeug.utils import secure_filename
import firebase_admin
from firebase_admin import credentials, firestore
import cloudinary
import cloudinary.uploader

app = Flask(__name__)

# ── SECRET KEY: must stay IDENTICAL across restarts or every session breaks ──
# Render/Heroku/etc restart the process often (sleep, redeploy, crash) — a
# random key generated at startup invalidates every logged-in user's cookie
# each time. Set FLASK_SECRET_KEY in your environment variables to a fixed
# random string. If not set, fall back to a generated key (sessions will not
# survive a restart in that case — set the env var to fix this permanently).
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "civision-default-key-change-in-production-2025")

# ── Session / cookie configuration ───────────────────────────────────────────
app.config.update(
    SESSION_COOKIE_NAME="civision_session",
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=True,
    PERMANENT_SESSION_LIFETIME=timedelta(days=15),
)

# ── Credentials from environment variables ──────────────────────────────────
FIREBASE_JSON   = os.environ.get("FIREBASE_JSON", "")
CLOUDINARY_NAME = os.environ.get("CLOUDINARY_NAME", "")
CLOUDINARY_KEY  = os.environ.get("CLOUDINARY_KEY", "")
CLOUDINARY_SEC  = os.environ.get("CLOUDINARY_SECRET", "")

MONTHLY_DUES    = 150
START_YEAR, START_MONTH = 2025, 12
CLUB_EMAIL      = "civisionsociety@gmail.com"
PAYMENT_PHONE   = "01838604302"
ADMIN_HASH      = hashlib.sha256("Zabed".encode()).hexdigest()
MASTER_HASH     = hashlib.sha256("Zabed1".encode()).hexdigest()

MONTHS_FULL  = {1:"January",2:"February",3:"March",4:"April",5:"May",6:"June",
                7:"July",8:"August",9:"September",10:"October",11:"November",12:"December"}
MONTHS_SHORT = {k:v[:3] for k,v in MONTHS_FULL.items()}

ALLOWED_EXT = {"png","jpg","jpeg","gif","webp"}

# ── Firebase init ────────────────────────────────────────────────────────────
_db = None
def get_db():
    global _db
    if _db is not None: return _db
    if not firebase_admin._apps:
        if FIREBASE_JSON:
            cred_dict = json.loads(FIREBASE_JSON)
            cred = credentials.Certificate(cred_dict)
        else:
            cred = credentials.ApplicationDefault()
        firebase_admin.initialize_app(cred)
    _db = firestore.client()
    return _db

# ── Cloudinary init ──────────────────────────────────────────────────────────
def init_cloudinary():
    cloudinary.config(cloud_name=CLOUDINARY_NAME,
                      api_key=CLOUDINARY_KEY,
                      api_secret=CLOUDINARY_SEC)

# ── Firestore helpers ────────────────────────────────────────────────────────
def col(name): return get_db().collection(name)

def get_all_members():
    docs = col("members").order_by("join_date").stream()
    return [{"id": d.id, **d.to_dict()} for d in docs]

def get_member_by_id(mid):
    doc = col("members").document(str(mid).lower()).get()
    if doc.exists: return {"id": doc.id, **doc.to_dict()}
    return None

def save_member(mid, data):
    col("members").document(str(mid).lower()).set(data)

def update_member(mid, data):
    col("members").document(str(mid).lower()).update(data)

def get_payments():
    docs = col("payments").stream()
    return [{"id": d.id, **d.to_dict()} for d in docs]

def get_payments_for_member(mid):
    docs = col("payments").where("member_id","==",str(mid).lower()).stream()
    return [{"id": d.id, **d.to_dict()} for d in docs]

def save_payment(pid, data):
    col("payments").document(pid).set(data)

def update_payment(pid, data):
    col("payments").document(pid).update(data)

def delete_payment(pid):
    col("payments").document(pid).delete()

def get_events():
    docs = col("events").order_by("event_date").stream()
    return [{"id": d.id, **d.to_dict()} for d in docs]

def get_gallery():
    docs = col("gallery").order_by("uploaded_at", direction=firestore.Query.DESCENDING).stream()
    return [{"id": d.id, **d.to_dict()} for d in docs]

def get_home_content():
    doc = col("settings").document("home").get()
    if doc.exists: return doc.to_dict()
    return {"about":"Civision Society is a dynamic student-led organization.",
            "announcements":"Welcome! Monthly dues are 150 BDT.",
            "notice":"General Body Meeting coming soon."}

def get_expenses():
    docs = col("expenses").order_by("date").stream()
    return [{"id": d.id, **d.to_dict()} for d in docs]

def get_admin_pw_hash():
    doc = col("settings").document("admin").get()
    if doc.exists: return doc.to_dict().get("password_hash", ADMIN_HASH)
    return ADMIN_HASH

def set_admin_pw_hash(h):
    col("settings").document("admin").set({"password_hash": h})

def get_saved_email_creds():
    doc = col("settings").document("email_creds").get()
    if doc.exists: return doc.to_dict()
    return {"sender_email":"", "sender_password":""}

def save_email_creds(email, password):
    col("settings").document("email_creds").set({"sender_email": email, "sender_password": password})

def get_notify_email():
    doc = col("settings").document("notify_email").get()
    if doc.exists: return doc.to_dict().get("email","")
    return ""

def save_notify_email(email):
    col("settings").document("notify_email").set({"email": email})

def send_admin_notification(subject, body_html, body_plain):
    """Send a notification to the admin's configured notify email using saved sender credentials."""
    notify_to = get_notify_email()
    if not notify_to: return False
    creds = get_saved_email_creds()
    se, sp = creds.get("sender_email",""), creds.get("sender_password","")
    if not se or not sp: return False
    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject; msg["From"] = se; msg["To"] = notify_to
        msg.attach(MIMEText(body_plain,"plain")); msg.attach(MIMEText(body_html,"html"))
        with smtplib.SMTP("smtp.gmail.com",587) as srv:
            srv.ehlo(); srv.starttls(); srv.login(se, sp)
            srv.sendmail(se, notify_to, msg.as_string())
        return True
    except Exception:
        return False

def check_admin_password(pw):
    h = hashlib.sha256(pw.encode()).hexdigest()
    return h == MASTER_HASH or h == get_admin_pw_hash()

# ── Session security helpers ─────────────────────────────────────────────────
def start_member_session(mid):
    """Log a member in: regenerate the session, bind it to a server-side
    revocable token, and make it persistent for 30 days on this device."""
    session.clear()                      # wipe any previous session data first
    session_token = secrets.token_hex(24)
    col("members").document(mid).update({"session_token": session_token})
    session["member_id"] = mid
    session["session_token"] = session_token
    session.permanent = True             # use PERMANENT_SESSION_LIFETIME (30 days)

def get_logged_in_member():
    """Returns the member dict if the current device has a valid, non-revoked
    session for an approved member. Returns None otherwise."""
    mid = session.get("member_id")
    token = session.get("session_token")
    if not mid or not token:
        return None
    member = get_member_by_id(mid)
    if not member or member.get("status") != "Approved":
        session.clear()
        return None
    if member.get("session_token") != token:
        # Token mismatch = session was revoked (e.g. password changed,
        # or admin force-logged-out this account) -> require fresh login.
        session.clear()
        return None
    return member

def end_member_session():
    mid = session.get("member_id")
    if mid:
        try:
            col("members").document(mid).update({"session_token": ""})
        except Exception:
            pass
    session.clear()

def start_admin_session():
    session.clear()
    session["admin"] = True
    session["admin_token"] = secrets.token_hex(24)
    session.permanent = True

def months_since_start():
    now = date.today()
    cur = date(START_YEAR, START_MONTH, 1)
    result = []
    while cur <= date(now.year, now.month, 1):
        result.append((cur.month, cur.year))
        m, y = cur.month, cur.year
        cur = date(y+(m==12), 1 if m==12 else m+1, 1)
    return result

def get_payment_status_for_member(mid):
    paid = {}
    for p in get_payments_for_member(mid):
        paid[(int(p.get("month",0)), int(p.get("year",0)))] = p
    result = []
    for mo, yr in months_since_start():
        p = paid.get((mo, yr))
        result.append({
            "month":        date(yr,mo,1).strftime("%B %Y"),
            "month_num":    mo, "year": yr,
            "status":       p["status"] if p else "Unpaid",
            "txn_id":       p.get("txn_id") if p else None,
            "submitted_at": p.get("submitted_at") if p else None,
            "pay_id":       p["id"] if p else None,
        })
    return result

def hash_pw(pw): return hashlib.sha256(pw.encode()).hexdigest()
def allowed_file(fn): return "." in fn and fn.rsplit(".",1)[1].lower() in ALLOWED_EXT

def upload_image(file_obj, folder="civision"):
    init_cloudinary()
    result = cloudinary.uploader.upload(file_obj, folder=folder)
    return result.get("secure_url","")

def admin_required(f):
    @wraps(f)
    def dec(*a, **kw):
        if not session.get("admin"): return redirect(url_for("admin_login"))
        return f(*a, **kw)
    return dec

def get_site_settings():
    doc = col("settings").document("site").get()
    if doc.exists: return doc.to_dict()
    return {"mobile_scale":"1.0", "bg_image":""}

def render(tmpl, **kw):
    kw.setdefault("now", datetime.now())
    kw.setdefault("club_email", CLUB_EMAIL)
    kw.setdefault("payment_phone", PAYMENT_PHONE)
    kw.setdefault("_title", "Civision Society")
    if "nav_logo" not in kw: kw["nav_logo"] = get_nav_logo()
    site = get_site_settings()
    kw.setdefault("mobile_scale", site.get("mobile_scale","1.0"))
    kw.setdefault("bg_image", site.get("bg_image",""))
    full = BASE.replace("{% block title %}Civision Society{% endblock %}", kw["_title"])
    full = full.replace("{% block body %}{% endblock %}", tmpl)
    return render_template_string(full, **kw)

def _tab(t): return redirect(url_for("admin_dashboard")+"#"+t)

# ── BASE TEMPLATE ────────────────────────────────────────────────────────────
BASE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" id="viewportMeta" content="width=device-width,initial-scale=1.0"/>
<title>{% block title %}Civision Society{% endblock %}</title>
<meta name="description" content="{{meta_description or 'Civision Society — a student-led club for civic education and purposeful networking. Join, log in, register, or pay your membership dues online.'}}"/>
<link rel="canonical" href="{{request.url_root.rstrip('/')}}{{request.path}}"/>
{% if nav_logo %}<link rel="icon" href="{{nav_logo}}"/><link rel="shortcut icon" href="{{nav_logo}}"/>{% endif %}

<!-- Open Graph -->
<meta property="og:title" content="{{_title or 'Civision Society'}}"/>
<meta property="og:description" content="{{meta_description or 'Civision Society — a student-led club for civic education and purposeful networking.'}}"/>
<meta property="og:type" content="website"/>
<meta property="og:url" content="{{request.url}}"/>
{% if nav_logo %}<meta property="og:image" content="{{nav_logo}}"/>{% endif %}
<meta property="og:site_name" content="Civision Society"/>

<!-- Twitter Card -->
<meta name="twitter:card" content="summary"/>
<meta name="twitter:title" content="Civision Society"/>
<meta name="twitter:description" content="{{meta_description or 'Civision Society — connected in purpose.'}}"/>
{% if nav_logo %}<meta name="twitter:image" content="{{nav_logo}}"/>{% endif %}

<!-- Structured data: Organization + SiteNavigationElement (helps Google show sitelinks) -->
<script type="application/ld+json">
{
  "@context": "https://schema.org",
  "@type": "Organization",
  "name": "Civision Society",
  "alternateName": "Connected in Purpose",
  "url": "{{request.url_root.rstrip('/')}}",
  {% if nav_logo %}"logo": "{{nav_logo}}",{% endif %}
  "sameAs": ["https://www.facebook.com/profile.php?id=61585468977506&mibextid=rS40aB7S9Ucbxw6v"]
}
</script>
<script type="application/ld+json">
{
  "@context": "https://schema.org",
  "@type": "WebSite",
  "name": "Civision Society",
  "url": "{{request.url_root.rstrip('/')}}"
}
</script>
<script type="application/ld+json">
{
  "@context": "https://schema.org",
  "@type": "ItemList",
  "itemListElement": [
    {"@type":"SiteNavigationElement","position":1,"name":"Home","url":"{{request.url_root.rstrip('/')}}/"},
    {"@type":"SiteNavigationElement","position":2,"name":"Profile","url":"{{request.url_root.rstrip('/')}}/login"},
    {"@type":"SiteNavigationElement","position":3,"name":"Registration Form","url":"{{request.url_root.rstrip('/')}}/join"},
    {"@type":"SiteNavigationElement","position":4,"name":"Payment","url":"{{request.url_root.rstrip('/')}}/payment"},
    {"@type":"SiteNavigationElement","position":5,"name":"Admin","url":"{{request.url_root.rstrip('/')}}/admin"}
  ]
}
</script>

<script>
(function(){
  var scale = "{{mobile_scale or '1.0'}}";
  if(window.innerWidth <= 768 && scale !== "1.0"){
    var vp = document.getElementById('viewportMeta');
    if(vp) vp.setAttribute('content','width=device-width,initial-scale='+scale);
  }
})();
</script>

<style>
:root{ {% if bg_image %}--bg-image: url('{{bg_image}}');{% endif %} }
@import url('https://fonts.googleapis.com/css2?family=Cormorant+Garamond:wght@700&family=Inter:wght@300;400;500;600&display=swap');
:root{--cream:#F5F2EC;--ink:#1A1A1A;--sl:#4A4A5A;--gold:#C9A84C;--wh:#FFF;
      --red:#C0392B;--grn:#27AE60;--blu:#2980B9;--bdr:#DDD8CE;--sh:0 2px 16px rgba(0,0,0,.08);}
*{box-sizing:border-box;margin:0;padding:0;}
html,body{min-height:100%;}
body{background:var(--cream) var(--bg-image, none) center/cover fixed no-repeat;color:var(--ink);font-family:'Inter',sans-serif;font-size:15px;line-height:1.7;
     display:flex;flex-direction:column;min-height:100vh;}
a{color:var(--gold);}a:hover{text-decoration:underline;}
nav{background:var(--ink);position:sticky;top:0;z-index:100;box-shadow:0 2px 12px rgba(0,0,0,.3);}
.ni{max-width:1200px;margin:auto;padding:0 20px;display:flex;align-items:center;justify-content:space-between;height:64px;gap:12px;}
.nb{display:flex;align-items:center;gap:10px;color:#fff;font-family:'Cormorant Garamond',serif;font-size:1.25rem;font-weight:700;letter-spacing:.04em;white-space:nowrap;text-decoration:none;}
.nb img{height:40px;width:40px;object-fit:contain;filter:brightness(0) invert(1);}
.nl{display:flex;align-items:center;gap:4px;flex-wrap:wrap;}
.nl a{color:#CCC;padding:6px 12px;border-radius:4px;font-size:.85rem;font-weight:500;transition:.2s;text-decoration:none;}
.nl a:hover,.nl a.act{background:var(--gold);color:var(--ink);}
.hb{display:none;flex-direction:column;gap:4px;cursor:pointer;padding:6px;}
.hb span{display:block;width:22px;height:2px;background:#fff;}
main{max-width:1200px;margin:24px auto 40px;padding:0 20px;flex:1 0 auto;width:100%;}
.bn{display:none;}
@media(max-width:768px){
  main{margin-bottom:90px;}
  .ph{text-align:center;padding:48px 20px 32px;border-bottom:1px solid var(--bdr);margin-bottom:40px;}
  .nl{display:none;}
  .bn{display:flex;position:fixed;bottom:0;left:0;right:0;background:var(--ink);
      box-shadow:0 -2px 12px rgba(0,0,0,.3);z-index:200;padding:6px 0 max(6px,env(safe-area-inset-bottom));}
  .bn a{flex:1;display:flex;flex-direction:column;align-items:center;justify-content:center;
        gap:3px;color:#999;text-decoration:none;font-size:.65rem;font-weight:500;padding:4px 2px;}
  .bn a .ic{font-size:1.25rem;line-height:1;}
  .bn a.act{color:var(--gold);}
}
.ph{text-align:center;padding:48px 20px 32px;border-bottom:1px solid var(--bdr);margin-bottom:40px;}
.ph h1{font-family:'Cormorant Garamond',serif;font-size:2.4rem;font-weight:700;margin-bottom:8px;}
.ph p{color:var(--sl);}
.card{background:var(--wh);border-radius:10px;box-shadow:var(--sh);padding:28px;margin-bottom:24px;}
.card h2{font-family:'Cormorant Garamond',serif;font-size:1.6rem;margin-bottom:16px;border-bottom:2px solid var(--gold);padding-bottom:8px;}
.card h3{font-size:1rem;font-weight:600;margin-bottom:12px;}
.fg{margin-bottom:16px;}
.fg label{display:block;font-size:.82rem;font-weight:600;color:var(--sl);margin-bottom:5px;text-transform:uppercase;letter-spacing:.05em;}
.fg input,.fg textarea,.fg select{width:100%;padding:10px 14px;border:1.5px solid var(--bdr);border-radius:6px;font-family:'Inter',sans-serif;font-size:.9rem;background:#FAFAF8;transition:.2s;}
.fg input:focus,.fg textarea:focus,.fg select:focus{outline:none;border-color:var(--gold);background:#fff;}
.fg textarea{resize:vertical;min-height:80px;}
.btn{display:inline-block;padding:10px 22px;border-radius:6px;border:none;font-size:.88rem;font-weight:600;cursor:pointer;transition:.2s;font-family:'Inter',sans-serif;text-decoration:none;}
.bp{background:var(--ink);color:#fff;}.bp:hover{background:#333;color:#fff;}
.bg{background:var(--gold);color:var(--ink);}.bg:hover{background:#b8943d;}
.br{background:var(--red);color:#fff;}.br:hover{background:#a93226;color:#fff;}
.bgrn{background:var(--grn);color:#fff;}.bgrn:hover{background:#219a52;color:#fff;}
.bbl{background:var(--blu);color:#fff;}.bbl:hover{background:#1a6a9a;color:#fff;}
.bsm{padding:6px 14px;font-size:.8rem;}.btn:hover{text-decoration:none;}
.al{padding:12px 16px;border-radius:6px;margin-bottom:16px;font-size:.9rem;}
.als{background:#d5f5e3;color:#1a5c38;border-left:4px solid var(--grn);}
.ale{background:#fadbd8;color:#7b241c;border-left:4px solid var(--red);}
.ali{background:#d6eaf8;color:#1a4a6b;border-left:4px solid var(--blu);}
.bdg{display:inline-block;padding:2px 10px;border-radius:20px;font-size:.75rem;font-weight:600;}
.bdgg{background:#d5f5e3;color:#1a5c38;}.bdgr{background:#fadbd8;color:#7b241c;}
.bdgy{background:#fef9e7;color:#7d6608;border:1px solid #f0d060;}.bdgb{background:#d6eaf8;color:#1a4a6b;}
.tw{overflow-x:auto;}
table{width:100%;border-collapse:collapse;font-size:.875rem;}
th{background:var(--ink);color:#fff;padding:10px 12px;text-align:left;font-size:.8rem;text-transform:uppercase;letter-spacing:.05em;}
td{padding:10px 12px;border-bottom:1px solid var(--bdr);vertical-align:middle;}
tr:hover td{background:#FAF9F6;}
.g2{display:grid;grid-template-columns:1fr 1fr;gap:24px;}
.g3{display:grid;grid-template-columns:repeat(3,1fr);gap:20px;}
.g4{display:grid;grid-template-columns:repeat(4,1fr);gap:16px;}
.gg{display:grid;grid-template-columns:repeat(auto-fill,minmax(200px,1fr));gap:12px;margin-top:16px;}
.gi{position:relative;border-radius:10px;overflow:hidden;aspect-ratio:1;cursor:pointer;}
.gi img{width:100%;height:100%;object-fit:cover;transition:.3s;}.gi:hover img{transform:scale(1.05);}
.gc{position:absolute;bottom:0;left:0;right:0;background:rgba(0,0,0,.6);color:#fff;font-size:.78rem;padding:6px 10px;transform:translateY(100%);transition:.3s;}
.gi:hover .gc{transform:translateY(0);}
.gg-page{display:none;}
.gg-page.act{display:grid;}
.gg-nav{display:flex;justify-content:center;align-items:center;gap:16px;margin-top:14px;}
.gg-nav button{background:var(--ink);color:#fff;border:none;width:36px;height:36px;border-radius:50%;
                font-size:1.1rem;cursor:pointer;display:flex;align-items:center;justify-content:center;}
.gg-nav button:disabled{opacity:.3;cursor:default;}
.gg-nav .gg-pageinfo{font-size:.82rem;color:var(--sl);font-weight:600;}
@media(max-width:768px){.gg{grid-template-columns:repeat(4,1fr);gap:6px;}}
.hero-slider{position:relative;border-radius:14px;overflow:hidden;box-shadow:var(--sh);margin-bottom:32px;background:var(--ink);}
.notice-marquee{background:var(--ink);overflow:hidden;white-space:nowrap;padding:8px 0;}
.notice-track{display:inline-block;animation:notice-scroll 22s linear infinite;}
.notice-track span{color:var(--gold);font-size:.85rem;font-weight:600;letter-spacing:.02em;}
@keyframes notice-scroll{0%{transform:translateX(100%);}100%{transform:translateX(-100%);}}
.hs-track{display:flex;transition:transform .5s ease;}
.hs-slide{flex:0 0 100%;position:relative;height:320px;}
.hs-slide img{width:100%;height:100%;object-fit:cover;}
.hs-slide .hs-overlay{position:absolute;inset:0;background:linear-gradient(to top,rgba(0,0,0,.75),rgba(0,0,0,.1) 50%);
                       display:flex;flex-direction:column;justify-content:flex-end;padding:24px;color:#fff;}
.hs-slide .hs-date{font-size:.72rem;color:var(--gold);font-weight:700;text-transform:uppercase;letter-spacing:.08em;margin-bottom:4px;}
.hs-slide h3{font-family:'Cormorant Garamond',serif;font-size:1.6rem;margin-bottom:4px;}
.hs-slide p{font-size:.85rem;color:rgba(255,255,255,.8);max-width:520px;}
.hs-dots{position:absolute;bottom:14px;right:18px;display:flex;gap:6px;z-index:5;}
.hs-dot{width:8px;height:8px;border-radius:50%;background:rgba(255,255,255,.4);cursor:pointer;transition:.2s;}
.hs-dot.act{background:var(--gold);width:20px;border-radius:4px;}
.hs-arrow{position:absolute;top:50%;transform:translateY(-50%);background:rgba(0,0,0,.4);color:#fff;border:none;
          width:36px;height:36px;border-radius:50%;font-size:1.1rem;cursor:pointer;z-index:5;display:flex;align-items:center;justify-content:center;}
.hs-arrow.l{left:12px;}.hs-arrow.r{right:12px;}
@media(max-width:768px){.hs-slide{height:220px;}.hs-slide h3{font-size:1.2rem;}}
.lb{display:none;position:fixed;inset:0;background:rgba(0,0,0,.9);z-index:9999;align-items:center;justify-content:center;}
.lb.open{display:flex;}.lb img{max-width:90vw;max-height:90vh;border-radius:8px;object-fit:contain;}
.lb-x{position:absolute;top:20px;right:28px;color:#fff;font-size:2.5rem;cursor:pointer;font-weight:300;}
.ev{background:#fff;border-radius:10px;box-shadow:var(--sh);overflow:hidden;}
.ev-b{height:180px;background:linear-gradient(135deg,var(--ink),#3a3a4a);display:flex;align-items:center;justify-content:center;overflow:hidden;}
.ev-b img{width:100%;height:100%;object-fit:cover;}
.ev-c{padding:20px;}.ev-c h3{font-family:'Cormorant Garamond',serif;font-size:1.3rem;margin-bottom:6px;}
.ev-c .dt{font-size:.78rem;color:var(--gold);font-weight:600;text-transform:uppercase;letter-spacing:.06em;margin-bottom:10px;}
.ev-c p{color:var(--sl);font-size:.9rem;}
.pr{display:flex;align-items:center;justify-content:space-between;padding:8px 12px;border-radius:6px;margin-bottom:6px;background:#FAFAF8;border:1px solid var(--bdr);}
.hh{text-align:center;padding:60px 20px 48px;}
.hh img{height:120px;margin-bottom:24px;}
.hh h1{font-family:'Cormorant Garamond',serif;font-size:3.2rem;font-weight:700;line-height:1.1;margin-bottom:8px;}
.hh .tg{color:var(--sl);font-size:.85rem;letter-spacing:.12em;text-transform:uppercase;margin-bottom:32px;}
.hh .cr{display:flex;gap:12px;justify-content:center;flex-wrap:wrap;}
.tr{display:flex;gap:6px;margin-bottom:24px;flex-wrap:wrap;overflow-x:auto;padding-bottom:4px;}
.tb{padding:8px 14px;border:1.5px solid var(--bdr);border-radius:6px;background:#fff;cursor:pointer;font-size:.82rem;font-weight:500;font-family:'Inter',sans-serif;transition:.2s;white-space:nowrap;}
.tb.act,.tb:hover{background:var(--ink);color:#fff;border-color:var(--ink);}
.tp{display:none;}.tp.act{display:block;}
hr.fn{border:none;border-top:1px solid var(--bdr);margin:24px 0;}
footer{background:var(--ink);color:#aaa;text-align:center;padding:28px 20px;font-size:.82rem;margin-top:60px;flex-shrink:0;}
footer span{color:var(--gold);}
footer a.fb-link{color:var(--gold);font-weight:600;text-decoration:none;}
footer a.fb-link:hover{text-decoration:underline;}
.sc{background:#fff;border-radius:10px;box-shadow:var(--sh);padding:20px;text-align:center;}
.sc .v{font-size:2rem;font-weight:700;}.sc .l{font-size:.78rem;color:var(--sl);text-transform:uppercase;letter-spacing:.06em;margin-top:4px;}
.nav_logo_url{display:none;}
@media(max-width:768px){.g2,.g3,.g4{grid-template-columns:1fr;}.hh h1{font-size:2.2rem;}.nl,.ns{display:none;}.hb{display:flex;}.ni{height:56px;}footer{margin-bottom:64px;}}
</style>
</head>
<body>
<nav>
  <div class="ni">
    <a href="/" class="nb">
      {% if nav_logo %}<img src="{{nav_logo}}" alt="Logo" onerror="this.style.display='none'"/>{% endif %}
      Civision Society
    </a>
    <div class="nl">
      <a href="/" class="{% if request.path=='/' %}act{% endif %}">Home</a>
      <a href="/join" class="{% if request.path=='/join' %}act{% endif %}">Registration Form</a>
      <a href="/payment" class="{% if request.path=='/payment' %}act{% endif %}">Payment</a>
      <a href="/login" class="{% if request.path=='/login' %}act{% endif %}">Profile</a>
      <a href="/admin" class="{% if '/admin' in request.path %}act{% endif %}">Admin</a>
    </div>
    <div class="hb" onclick="document.getElementById('mm').classList.toggle('open')" style="display:none;">
      <span></span><span></span><span></span>
    </div>
  </div>
</nav>
<div class="bn">
  <a href="/" class="{% if request.path=='/' %}act{% endif %}"><span class="ic">🏠</span><span>Home</span></a>
  <a href="/login" class="{% if request.path=='/login' or request.path=='/member/profile' %}act{% endif %}"><span class="ic">👤</span><span>Profile</span></a>
  <a href="/join" class="{% if request.path=='/join' %}act{% endif %}"><span class="ic">📝</span><span>Register</span></a>
  <a href="/payment" class="{% if request.path=='/payment' %}act{% endif %}"><span class="ic">💳</span><span>Payment</span></a>
  <a href="/admin" class="{% if '/admin' in request.path %}act{% endif %}"><span class="ic">⚙️</span><span>Admin</span></a>
</div>
{% with msgs=get_flashed_messages(with_categories=true) %}
{% if msgs %}<div style="max-width:1200px;margin:16px auto 0;padding:0 20px;">
  {% for cat,msg in msgs %}<div class="al al{{cat}}">{{msg}}</div>{% endfor %}
</div>{% endif %}{% endwith %}
<main>{% block body %}{% endblock %}</main>
<div class="lb" id="lb" onclick="closeLB()">
  <span class="lb-x">×</span>
  <img id="lbimg" src="" alt=""/>
</div>
<footer><a class="fb-link" href="https://www.facebook.com/profile.php?id=61585468977506&mibextid=rS40aB7S9Ucbxw6v" target="_blank" rel="noopener">Civision Society</a> — "Connected in Purpose"<br/>
{{club_email}} &nbsp;|&nbsp; bKash/Nagad: {{payment_phone}}</footer>
<script>
function openLB(s){document.getElementById('lbimg').src=s;document.getElementById('lb').classList.add('open');}
function closeLB(){document.getElementById('lb').classList.remove('open');}
function showTab(id,btn){
  document.querySelectorAll('.tp').forEach(p=>p.classList.remove('act'));
  document.querySelectorAll('.tb').forEach(b=>b.classList.remove('act'));
  var p=document.getElementById(id);if(p)p.classList.add('act');
  if(btn)btn.classList.add('act');
  try{sessionStorage.setItem('atab',id);}catch(e){}
}
window.addEventListener('DOMContentLoaded',function(){
  var t=null;try{t=sessionStorage.getItem('atab');}catch(e){}
  var h=window.location.hash.replace('#','');
  var id=h||t;
  if(id){
    var p=document.getElementById(id);
    if(p){
      document.querySelectorAll('.tp').forEach(x=>x.classList.remove('act'));
      document.querySelectorAll('.tb').forEach(x=>x.classList.remove('act'));
      p.classList.add('act');
      var b=document.querySelector('[data-tab="'+id+'"]');
      if(b)b.classList.add('act');
    }
  }
});
</script>
</body></html>"""

def get_nav_logo():
    doc = col("settings").document("logo").get()
    if doc.exists: return doc.to_dict().get("url","")
    return ""

# ── HOME ──────────────────────────────────────────────────────────────────────
@app.route("/")
def home():
    content = get_home_content()
    gallery = get_gallery()
    evs = get_events()
    nav_logo = get_nav_logo()
    tmpl = r"""
{% if notice %}
<div class="notice-marquee">
  <div class="notice-track"><span>📢 &nbsp;{{notice}}&nbsp; &nbsp; &nbsp; 📢 &nbsp;{{notice}}</span></div>
</div>
{% endif %}
<div class="hh" style="padding:32px 20px 28px;">
  {% if nav_logo %}<img src="{{nav_logo}}" alt="Civision" style="filter:drop-shadow(0 4px 12px rgba(0,0,0,.12));height:90px;" onerror="this.style.display='none'"/>{% endif %}
  <h1 style="font-size:2.4rem;">Civision Society</h1>
  <div class="tg">"Connected in Purpose"</div>
  <div class="cr">
    <a href="/join" class="btn bg">Become a Member</a>
    <a href="/payment" class="btn bp">Pay Dues</a>
  </div>
</div>

{% if events %}
<div class="hero-slider" id="heroSlider">
  <div class="hs-track" id="hsTrack">
    {% for ev in events %}
    <div class="hs-slide">
      {% if ev.image_url %}<img src="{{ev.image_url}}" alt="{{ev.title}}"/>
      {% else %}<div style="width:100%;height:100%;background:linear-gradient(135deg,#1a1a1a,#3a3a4a);"></div>{% endif %}
      <div class="hs-overlay">
        <div class="hs-date">{{ev.event_date}}</div>
        <h3>{{ev.title}}</h3>
        <p>{{ev.description}}</p>
      </div>
    </div>
    {% endfor %}
  </div>
  {% if events|length > 1 %}
  <button class="hs-arrow l" onclick="hsMove(-1)">‹</button>
  <button class="hs-arrow r" onclick="hsMove(1)">›</button>
  <div class="hs-dots" id="hsDots">
    {% for ev in events %}<div class="hs-dot" onclick="hsGoTo({{loop.index0}})"></div>{% endfor %}
  </div>
  {% endif %}
</div>
<script>
(function(){
  var idx=0; var track=document.getElementById('hsTrack');
  var dots=document.querySelectorAll('#hsDots .hs-dot');
  var total=track ? track.children.length : 0;
  function render(){
    if(!track)return;
    track.style.transform='translateX(-'+(idx*100)+'%)';
    dots.forEach(function(d,i){d.className='hs-dot'+(i==idx?' act':'');});
  }
  window.hsMove=function(dir){idx=(idx+dir+total)%total;render();};
  window.hsGoTo=function(i){idx=i;render();};
  render();
  if(total>1){setInterval(function(){hsMove(1);},5000);}
})();
</script>
{% endif %}

<div class="g2" style="margin-bottom:24px;">
  <div class="card"><h2>About Us</h2><p style="color:var(--sl);line-height:1.8;">{{about}}</p></div>
  <div class="card"><h2>Announcements</h2><p style="color:var(--sl);line-height:1.8;">{{ann}}</p></div>
</div>

{% if gallery %}
<div class="card">
  <h2>Gallery</h2>
  <div id="galWrap"></div>
  <div class="gg-nav" id="galNav" style="display:none;">
    <button id="galPrev" onclick="galMove(-1)">‹</button>
    <span class="gg-pageinfo" id="galInfo"></span>
    <button id="galNext" onclick="galMove(1)">›</button>
  </div>
</div>
<script>
(function(){
  var images = {{ gallery_json|safe }};
  var perPage = window.innerWidth <= 768 ? 16 : 12;
  var page = 0;
  var totalPages = Math.ceil(images.length / perPage);
  function render(){
    var start = page*perPage, slice = images.slice(start, start+perPage);
    var html = '<div class="gg">';
    slice.forEach(function(img){
      html += '<div class="gi" onclick="openLB(\''+img.url+'\')"><img src="'+img.url+'" alt="'+(img.caption||'')+'"/>'+
              (img.caption ? '<div class="gc">'+img.caption+'</div>' : '')+'</div>';
    });
    html += '</div>';
    document.getElementById('galWrap').innerHTML = html;
    if(totalPages>1){
      document.getElementById('galNav').style.display='flex';
      document.getElementById('galInfo').textContent=(page+1)+' / '+totalPages;
      document.getElementById('galPrev').disabled = page===0;
      document.getElementById('galNext').disabled = page>=totalPages-1;
    }
  }
  window.galMove=function(dir){
    page=Math.max(0,Math.min(totalPages-1,page+dir));
    render();
  };
  render();
})();
</script>
{% endif %}"""
    gallery_json = json.dumps([{"url":g.get("url",""),"caption":g.get("caption","")} for g in gallery])
    return render(tmpl, _title="Civision Society — Home",
                  about=content.get("about",""), ann=content.get("announcements",""),
                  notice=content.get("notice",""), gallery=gallery, events=evs,
                  gallery_json=gallery_json, nav_logo=nav_logo,
                  meta_description="Civision Society official website — student-led club for civic education and purposeful networking. View events, join as a member, log in, or pay your dues online.")

# ── JOIN ──────────────────────────────────────────────────────────────────────
@app.route("/join", methods=["GET","POST"])
def join():
    if request.method == "POST":
        mid   = request.form.get("member_id","").strip().lower()
        name  = request.form.get("name","").strip()
        phone = request.form.get("phone","").strip()
        email = request.form.get("email","").strip()
        pw    = request.form.get("password","").strip()
        uni   = request.form.get("university","").strip()
        dept  = request.form.get("department","").strip()
        blood = request.form.get("blood_group","").strip()
        dob   = request.form.get("dob","").strip()
        addr  = request.form.get("address","").strip()
        if not all([mid,name,phone,email,pw,uni,dept]):
            flash("All required fields must be filled.","e"); return redirect(url_for("join"))
        if get_member_by_id(mid):
            flash("That Member ID is already taken.","e"); return redirect(url_for("join"))
        save_member(mid, {"id":mid,"name":name,"phone":phone,"email":email,
                          "password_hash":hash_pw(pw),"university":uni,"department":dept,
                          "blood_group":blood,"dob":dob,"address":addr,
                          "status":"Pending","join_date":datetime.now().strftime("%Y-%m-%d")})
        try:
            subj = f"New Registration — {name} ({mid})"
            plain = (f"A new membership registration has been submitted.\n\n"
                     f"Name: {name}\nMember ID: {mid}\nPhone: {phone}\nEmail: {email}\n"
                     f"University: {uni}\nDepartment: {dept}\n\n"
                     f"Please review and approve/reject this application in the Admin Dashboard.")
            html = (f'<div style="font-family:sans-serif;max-width:520px;margin:auto;padding:20px;">'
                    f'<h2 style="font-family:Georgia,serif;color:#1a1a1a;border-bottom:2px solid #C9A84C;padding-bottom:8px;">New Registration</h2>'
                    f'<p>A new membership application has been submitted:</p>'
                    f'<div style="background:#f9f9f7;border-radius:8px;padding:16px;margin:14px 0;border-left:4px solid #C9A84C;">'
                    f'<p><strong>Name:</strong> {name}</p><p><strong>Member ID:</strong> {mid}</p>'
                    f'<p><strong>Phone:</strong> {phone}</p><p><strong>Email:</strong> {email}</p>'
                    f'<p><strong>University:</strong> {uni}</p><p><strong>Department:</strong> {dept}</p></div>'
                    f'<p>Please review and approve/reject this application in the Admin Dashboard.</p></div>')
            send_admin_notification(subj, html, plain)
        except Exception:
            pass
        flash("Application submitted! Wait for Admin approval.","s")
        return redirect(url_for("home"))
    tmpl = r"""
<div class="ph"><h1>Join Civision Society</h1><p>Fill in your details. Admin will review before activation.</p></div>
<div style="max-width:640px;margin:auto;">
  <div class="card"><h2>Membership Application</h2>
    <form method="post">
      <div class="g2">
        <div class="fg"><label>Full Name *</label><input name="name" required placeholder="Your full name"/></div>
        <div class="fg"><label>Member ID *</label><input name="member_id" required placeholder="e.g. CS-001"/></div>
      </div>
      <div class="g2">
        <div class="fg"><label>Phone *</label><input name="phone" required placeholder="01XXXXXXXXX"/></div>
        <div class="fg"><label>Email *</label><input name="email" type="email" required placeholder="you@example.com"/></div>
      </div>
      <div class="g2">
        <div class="fg"><label>University *</label><input name="university" required placeholder="Your university"/></div>
        <div class="fg"><label>Department *</label><input name="department" required placeholder="e.g. CSE, EEE"/></div>
      </div>
      <div class="g2">
        <div class="fg"><label>Blood Group</label>
          <select name="blood_group"><option value="">-- Select --</option>
            {% for bg in ['A+','A-','B+','B-','AB+','AB-','O+','O-'] %}<option>{{bg}}</option>{% endfor %}
          </select>
        </div>
        <div class="fg"><label>Date of Birth</label><input type="date" name="dob"/></div>
      </div>
      <div class="fg"><label>Password *</label><input name="password" type="password" required placeholder="Strong password"/></div>
      <div class="fg"><label>Address (Optional)</label><input name="address" placeholder="Your address"/></div>
      <button class="btn bg" type="submit" style="width:100%;padding:12px;">Submit Application</button>
    </form>
  </div>
</div>"""
    return render(tmpl, _title="Registration Form — Civision Society", nav_logo=get_nav_logo(),
                  meta_description="Register as a new member of Civision Society. Fill out the registration form to join our community.")

# ── MEMBER LOGIN & PROFILE ───────────────────────────────────────────────────
@app.route("/login", methods=["GET","POST"])
def member_login():
    if request.method == "POST":
        mid = request.form.get("member_id","").strip().lower()
        pw  = request.form.get("password","").strip()
        member = get_member_by_id(mid)
        if not member:
            flash("Member ID not found.","e"); return redirect(url_for("member_login"))
        if member.get("status") != "Approved":
            flash("Your account is pending admin approval.","i"); return redirect(url_for("member_login"))
        if member.get("password_hash") != hash_pw(pw):
            flash("Incorrect password.","e"); return redirect(url_for("member_login"))
        start_member_session(mid)
        return redirect(url_for("member_profile"))
    tmpl = r"""
<div style="max-width:400px;margin:80px auto;">
  <div class="card" style="text-align:center;">
    {% if nav_logo %}<img src="{{nav_logo}}" style="height:64px;margin-bottom:16px;filter:brightness(0) invert(0);" onerror="this.style.display='none'"/>{% endif %}
    <h2 style="border:none;">Profile</h2>
    <p style="color:var(--sl);margin-bottom:20px;font-size:.9rem;">Civision Society — Member Portal</p>
    <form method="post">
      <div class="fg"><label>Member ID</label><input name="member_id" required placeholder="e.g. CS-001" autofocus/></div>
      <div class="fg"><label>Password</label><input name="password" type="password" required placeholder="Your password"/></div>
      <button class="btn bg" type="submit" style="width:100%;">Log In</button>
    </form>
    <p style="text-align:center;margin-top:16px;font-size:.85rem;color:var(--sl);">
      Not a member yet? <a href="/join">Register here</a>
    </p>
  </div>
</div>"""
    return render(tmpl, _title="Profile — Civision Society", nav_logo=get_nav_logo(),
                  meta_description="Member login for Civision Society. Log in with your Member ID and password to view your profile and payment status.")

@app.route("/member/logout")
def member_logout():
    end_member_session()
    return redirect(url_for("home"))

@app.route("/member/profile")
def member_profile():
    member = get_logged_in_member()
    if not member:
        flash("Please log in to view your profile.","i")
        return redirect(url_for("member_login"))
    mid = member["id"]
    all_payments = get_payments()
    all_expenses = get_expenses()
    total_collected = sum(MONTHLY_DUES for p in all_payments if p.get("status")=="Verified")
    total_expenses  = sum(float(e.get("amount",0) or 0) for e in all_expenses)
    club_balance    = total_collected - total_expenses
    tmpl = r"""
<div style="max-width:680px;margin:40px auto;">
  <div class="card">
    <div style="display:flex;align-items:center;gap:20px;margin-bottom:16px;flex-wrap:wrap;">
      {% if member.photo_url %}
        <img src="{{member.photo_url}}" alt="{{member.name}}" style="width:90px;height:90px;border-radius:50%;object-fit:cover;border:3px solid var(--gold);"/>
      {% else %}
        <div style="width:90px;height:90px;border-radius:50%;background:var(--bdr);display:flex;align-items:center;justify-content:center;font-size:2rem;color:var(--sl);">👤</div>
      {% endif %}
      <div>
        <h2 style="margin:0;border:none;">{{member.name}}</h2>
        <a href="/member/logout" style="font-size:.82rem;">Log Out</a>
      </div>
    </div>
    <div class="g2" style="margin-bottom:16px;">
      <div><strong>ID:</strong> {{member.id}}</div>
      <div><strong>Phone:</strong> {{member.phone}}</div>
      <div><strong>University:</strong> {{member.university}}</div>
      <div><strong>Department:</strong> {{member.department or '—'}}</div>
      <div><strong>Blood Group:</strong> {{member.blood_group or '—'}}</div>
      <div><strong>Date of Birth:</strong> {{member.dob or '—'}}</div>
      <div><strong>Member Since:</strong> {{member.join_date}}</div>
    </div>
  </div>
  <div class="card"><h2>📷 Update Profile Photo</h2>
    <form method="post" action="/member/photo/upload" enctype="multipart/form-data">
      <div class="fg"><label>Current Password *</label><input name="current_password" type="password" required placeholder="Confirm your password"/></div>
      <div class="fg"><label>Choose Photo *</label><input type="file" name="photo" accept="image/*" required/></div>
      <button class="btn bg" type="submit">Upload Photo</button>
    </form>
  </div>
  <div class="card"><h2>Update Profile Details</h2>
    <form method="post" action="/member/update">
      <div class="fg"><label>Current Password *</label><input name="current_password" type="password" required/></div>
      <hr class="fn"/>
      <div class="g2">
        <div class="fg"><label>Phone</label><input name="new_phone" value="{{member.phone or ''}}"/></div>
        <div class="fg"><label>University</label><input name="new_university" value="{{member.university or ''}}"/></div>
        <div class="fg"><label>Department</label><input name="new_department" value="{{member.department or ''}}"/></div>
        <div class="fg"><label>Address</label><input name="new_address" value="{{member.address or ''}}"/></div>
      </div>
      <button class="btn bp" type="submit">Save Details</button>
    </form>
  </div>
  <div class="card"><h2>🔐 Change Password</h2>
    <form method="post" action="/member/update">
      <div class="fg"><label>Current Password *</label><input name="current_password" type="password" required placeholder="Required to confirm any change"/></div>
      <div class="g2">
        <div class="fg"><label>New Password</label><input name="new_password" type="password" placeholder="Leave blank to keep current"/></div>
        <div class="fg"><label>Confirm New Password</label><input name="confirm_password" type="password"/></div>
      </div>
      <button class="btn bp" type="submit">Change Password</button>
    </form>
  </div>
  <div class="card" style="background:linear-gradient(135deg,#145a32,#1e8449);color:#fff;text-align:center;">
    <h2 style="color:#fff;border-color:rgba(255,255,255,.3);">Current Club Balance</h2>
    <div style="font-size:2.2rem;font-weight:700;">{{club_balance|int}} ৳</div>
    <div style="color:rgba(255,255,255,.7);font-size:.85rem;margin-top:4px;">Total collected dues minus club expenses</div>
  </div>
</div>"""
    return render(tmpl, _title="My Profile — Civision Society", member=member,
                  club_balance=club_balance, nav_logo=get_nav_logo())

@app.route("/member/photo/upload", methods=["POST"])
def member_photo_upload():
    member = get_logged_in_member()
    if not member: flash("Please log in first.","e"); return redirect(url_for("member_login"))
    mid = member["id"]
    cur_pw = request.form.get("current_password","").strip()
    if member.get("password_hash") != hash_pw(cur_pw): flash("Incorrect password.","e"); return redirect(url_for("member_profile"))
    if "photo" not in request.files: flash("No photo selected.","e"); return redirect(url_for("member_profile"))
    f = request.files["photo"]
    if not f or not f.filename or not allowed_file(f.filename):
        flash("Please select a valid image file.","e"); return redirect(url_for("member_profile"))
    url = upload_image(f, folder="civision/member_photos")
    update_member(mid, {"photo_url": url})
    flash("Profile photo updated!","s")
    return redirect(url_for("member_profile"))

@app.route("/member/update", methods=["POST"])
def member_update():
    member = get_logged_in_member()
    if not member: flash("Please log in first.","e"); return redirect(url_for("member_login"))
    mid = member["id"]
    cur_pw = request.form.get("current_password","").strip()
    if member.get("password_hash") != hash_pw(cur_pw): flash("Incorrect password.","e"); return redirect(url_for("member_profile"))
    new_pw = request.form.get("new_password","").strip()
    conf   = request.form.get("confirm_password","").strip()
    if new_pw and new_pw != conf: flash("Passwords do not match.","e"); return redirect(url_for("member_profile"))
    updates = {}
    for field, form_key in [("phone","new_phone"),("university","new_university"),("department","new_department"),("address","new_address")]:
        val = request.form.get(form_key,"").strip()
        if val: updates[field] = val
    if new_pw:
        updates["password_hash"] = hash_pw(new_pw)
        # Changing password invalidates this session's token too; re-issue
        # a fresh one immediately so the user isn't logged out on this device.
        new_token = secrets.token_hex(24)
        updates["session_token"] = new_token
        session["session_token"] = new_token
    if updates: update_member(mid, updates)
    flash("Profile updated!","s")
    return redirect(url_for("member_profile"))

# ── PAYMENT ──────────────────────────────────────────────────────────────────
@app.route("/payment", methods=["GET","POST"])
def payment():
    member = None; pay_status = []
    q = request.args.get("mid","").strip().lower()
    if q:
        member = get_member_by_id(q)
        if member and member.get("status") == "Approved":
            pay_status = get_payment_status_for_member(q)
    if request.method == "POST":
        action = request.form.get("action","")
        if action == "lookup":
            return redirect(url_for("payment", mid=request.form.get("member_id","").strip().lower()))
        if action == "submit":
            mid   = request.form.get("member_id","").strip().lower()
            month = request.form.get("month","")
            year  = request.form.get("year","")
            txn   = request.form.get("txn_id","").strip()
            if not all([mid,month,year,txn]): flash("All fields required.","e"); return redirect(url_for("payment",mid=mid))
            m = get_member_by_id(mid)
            if not m or m.get("status") != "Approved": flash("Member not found or not approved.","e"); return redirect(url_for("payment",mid=mid))
            pid = f"{mid}_{month}_{year}"
            existing = col("payments").document(pid).get()
            if existing.exists: flash("Record exists for this month. Contact admin.","e"); return redirect(url_for("payment",mid=mid))
            save_payment(pid, {"member_id":mid,"month":int(month),"year":int(year),
                               "status":"Pending","txn_id":txn,
                               "submitted_at":datetime.now().strftime("%Y-%m-%d %H:%M"),"verified_at":""})
            try:
                month_label = date(int(year), int(month), 1).strftime("%B %Y")
                subj = f"New Payment Submitted — {m.get('name')} ({mid})"
                plain = (f"A new payment has been submitted for verification.\n\n"
                         f"Member: {m.get('name')} ({mid})\nMonth: {month_label}\n"
                         f"Amount: {MONTHLY_DUES} BDT\nTransaction ID: {txn}\n\n"
                         f"Please verify this payment in the Admin Dashboard.")
                html = (f'<div style="font-family:sans-serif;max-width:520px;margin:auto;padding:20px;">'
                        f'<h2 style="font-family:Georgia,serif;color:#1a1a1a;border-bottom:2px solid #C9A84C;padding-bottom:8px;">New Payment Submitted</h2>'
                        f'<div style="background:#f9f9f7;border-radius:8px;padding:16px;margin:14px 0;border-left:4px solid #C9A84C;">'
                        f'<p><strong>Member:</strong> {m.get("name")} ({mid})</p>'
                        f'<p><strong>Month:</strong> {month_label}</p>'
                        f'<p><strong>Amount:</strong> {MONTHLY_DUES} BDT</p>'
                        f'<p><strong>Transaction ID:</strong> {txn}</p></div>'
                        f'<p>Please verify this payment in the Admin Dashboard.</p></div>')
                send_admin_notification(subj, html, plain)
            except Exception:
                pass
            flash("Payment submitted! Admin will verify shortly.","s")
            return redirect(url_for("payment",mid=mid))
    all_months = months_since_start()
    unpaid = [p for p in pay_status if p["status"]=="Unpaid"]
    month_options = [(p["month_num"],p["year"]) for p in unpaid] if unpaid else all_months
    tmpl = r"""
<div class="ph"><h1>Payment Desk</h1><p>Look up your dues and submit your bKash / Nagad transaction.</p></div>
<div style="max-width:640px;margin:auto;">
  <div class="card" style="background:linear-gradient(135deg,var(--ink),#2c3e50);color:#fff;text-align:center;">
    <h2 style="color:#fff;border-color:#444;">Pay Via bKash / Nagad</h2>
    <div style="font-size:2rem;font-weight:700;color:var(--gold);">{{payment_phone}}</div>
    <div style="color:#aaa;margin-top:4px;">150 ৳ per month — Personal bKash / Nagad</div>
  </div>
  <div class="card"><h2>Look Up Your Account</h2>
    <form method="post"><input type="hidden" name="action" value="lookup"/>
      <div style="display:flex;gap:10px;">
        <input style="flex:1;padding:10px 14px;border:1.5px solid var(--bdr);border-radius:6px;"
               name="member_id" placeholder="Enter your Member ID…" value="{{q}}"/>
        <button class="btn bg" type="submit">Lookup</button>
      </div>
    </form>
  </div>
  {% if member and member.status=='Approved' %}
  <div class="card"><h2>Submit Transaction — {{member.name}}</h2>
    <form method="post"><input type="hidden" name="action" value="submit"/>
      <input type="hidden" name="member_id" value="{{member.id}}"/>
      <div class="g2">
        <div class="fg"><label>Month &amp; Year *</label>
          <select name="amo" id="psel" onchange="sp(this)">
            {% for mo,yr in month_options %}<option value="{{mo}}|{{yr}}">{{mfull[mo]}} {{yr}}</option>{% endfor %}
          </select>
          <input type="hidden" name="month" id="pmo" value="{{month_options[0][0] if month_options else ''}}"/>
          <input type="hidden" name="year" id="pyr" value="{{month_options[0][1] if month_options else ''}}"/>
          <script>function sp(s){var p=s.value.split('|');document.getElementById('pmo').value=p[0];document.getElementById('pyr').value=p[1];}
          var ps=document.getElementById('psel');if(ps)sp(ps);</script>
        </div>
        <div class="fg"><label>Transaction ID *</label><input name="txn_id" required placeholder="e.g. AB1234567890"/></div>
      </div>
      <button class="btn bg" type="submit" style="width:100%;">Submit for Verification</button>
    </form>
  </div>
  <div class="card"><h2>Your Payment Status</h2>
    {% for p in pay_status %}
    <div class="pr"><span style="font-weight:500;">{{p.month}}</span>
      {% if p.status=='Verified' %}<span class="bdg bdgg">✓ Paid</span>
      {% elif p.status=='Pending' %}<span class="bdg bdgy">⏳ Pending</span>
      {% else %}<span class="bdg bdgr">✗ Unpaid — 150 ৳</span>{% endif %}
    </div>{% endfor %}
  </div>
  {% elif q and not member %}<div class="al ale">Member ID "{{q}}" not found.</div>
  {% elif q and member and member.status!='Approved' %}<div class="al ali">Account pending approval.</div>
  {% endif %}
</div>"""
    return render(tmpl, _title="Payment — Civision Society",
                  q=q, member=member, pay_status=pay_status,
                  month_options=month_options, mfull=MONTHS_FULL, nav_logo=get_nav_logo(),
                  meta_description="Pay your Civision Society membership dues online via bKash or Nagad. Submit your transaction ID for verification.")

# ── ADMIN LOGIN ───────────────────────────────────────────────────────────────
@app.route("/admin", methods=["GET","POST"])
def admin_login():
    if session.get("admin"): return redirect(url_for("admin_dashboard"))
    if request.method == "POST":
        if check_admin_password(request.form.get("password","")):
            start_admin_session()
            flash("Welcome to the Admin Portal.","s")
            return redirect(url_for("admin_dashboard"))
        flash("Incorrect password.","e")
    tmpl = r"""
<div style="max-width:400px;margin:80px auto;">
  <div class="card" style="text-align:center;">
    {% if nav_logo %}<img src="{{nav_logo}}" style="height:64px;margin-bottom:16px;filter:brightness(0) invert(0);" onerror="this.style.display='none'"/>{% endif %}
    <h2 style="border:none;">Admin Portal</h2>
    <p style="color:var(--sl);margin-bottom:20px;font-size:.9rem;">Civision Society Administration</p>
    <form method="post">
      <div class="fg"><label>Password</label><input name="password" type="password" required autofocus/></div>
      <button class="btn bg" type="submit" style="width:100%;">Access Terminal</button>
    </form>
  </div>
</div>"""
    return render(tmpl, _title="Admin Login", nav_logo=get_nav_logo())

@app.route("/admin/logout")
def admin_logout():
    session.clear()
    return redirect(url_for("home"))

# ── ADMIN DASHBOARD ───────────────────────────────────────────────────────────
@app.route("/admin/dashboard")
@admin_required
def admin_dashboard():
    members  = get_all_members()
    pending  = [m for m in members if m.get("status")=="Pending"]
    approved = [m for m in members if m.get("status")=="Approved"]
    payments = get_payments()
    pending_p= [p for p in payments if p.get("status")=="Pending"]
    events   = get_events()
    gallery  = get_gallery()
    content  = get_home_content()
    expenses = get_expenses()
    month_options = months_since_start()
    pay_lookup = {}
    for p in payments:
        pay_lookup[(str(p.get("member_id","")).lower(), int(p.get("month",0)), int(p.get("year",0)))] = p
    total_col=total_ver=total_pend_amt=total_due_amt=0
    member_summary = []
    for m in approved:
        cells=[]; paid_c=due_c=paid_amt=0; unpaid_months=[]
        for (mo,yr) in month_options:
            p = pay_lookup.get((str(m.get("id","")).lower(), mo, yr))
            st = p["status"] if p else "Due"
            if st=="Verified":
                cells.append("V"); paid_c+=1; paid_amt+=MONTHLY_DUES; total_col+=MONTHLY_DUES; total_ver+=1
            elif st=="Pending":
                cells.append("P"); total_pend_amt+=MONTHLY_DUES
                unpaid_months.append(date(yr,mo,1).strftime("%B %Y"))
            else:
                cells.append("-"); due_c+=1; total_due_amt+=MONTHLY_DUES
                unpaid_months.append(date(yr,mo,1).strftime("%B %Y"))
        member_summary.append({"id":m.get("id"),"name":m.get("name"),"email":m.get("email") or "",
                                "cells":cells,"paid_count":paid_c,"due_count":due_c,
                                "paid_amount":paid_amt,"unpaid_months":unpaid_months})
    total_exp = sum(float(e.get("amount",0) or 0) for e in expenses)
    cur_bal   = total_col - total_exp
    summary_months = [f"{MONTHS_SHORT[m][:3]}{str(y)[-2:]}" for m,y in month_options]
    nav_logo = get_nav_logo()
    tmpl = r"""
<div style="display:flex;align-items:flex-start;justify-content:space-between;gap:12px;flex-wrap:wrap;margin-bottom:20px;">
<div class="tr" style="margin-bottom:0;flex:1;">
  <button class="tb act" data-tab="tab-apr" onclick="showTab('tab-apr',this)">Approvals <span class="bdg bdgr">{{pending|length}}</span></button>
  <button class="tb" data-tab="tab-pay" onclick="showTab('tab-pay',this)">Payments</button>
  <button class="tb" data-tab="tab-mem" onclick="showTab('tab-mem',this)">All Members</button>
  <button class="tb" data-tab="tab-sum" onclick="showTab('tab-sum',this)">Summary</button>
  <button class="tb" data-tab="tab-db" onclick="showTab('tab-db',this)">🗄 DB Editor</button>
  <button class="tb" data-tab="tab-cnt" onclick="showTab('tab-cnt',this)">Content</button>
  <button class="tb" data-tab="tab-gal" onclick="showTab('tab-gal',this)">Gallery</button>
  <button class="tb" data-tab="tab-ev" onclick="showTab('tab-ev',this)">Events</button>
  <button class="tb" data-tab="tab-eml" onclick="showTab('tab-eml',this)">Email</button>
  <button class="tb" data-tab="tab-logo" onclick="showTab('tab-logo',this)">Logo & Settings</button>
</div>
<a href="/admin/logout" class="btn br bsm" style="flex-shrink:0;">Log Out</a>
</div>

<!-- APPROVALS -->
<div id="tab-apr" class="tp act"><div class="card"><h2>Pending Approvals</h2>
{% if not pending %}<p style="color:var(--sl);">No pending applications.</p>{% else %}
<div class="tw"><table>
<thead><tr><th>ID</th><th>Name</th><th>Phone</th><th>Email</th><th>University</th><th>Dept</th><th>Blood</th><th>DOB</th><th>Applied</th><th>Actions</th></tr></thead>
<tbody>{% for m in pending %}<tr>
  <td>{{m.id}}</td><td>{{m.name}}</td><td>{{m.phone}}</td><td>{{m.email}}</td>
  <td>{{m.university}}</td><td>{{m.department or '—'}}</td><td>{{m.blood_group or '—'}}</td>
  <td>{{m.dob or '—'}}</td><td>{{m.join_date}}</td>
  <td style="display:flex;gap:6px;">
    <form method="post" action="/admin/approve"><input type="hidden" name="member_id" value="{{m.id}}"/><button class="btn bgrn bsm">✓ Approve</button></form>
    <form method="post" action="/admin/reject"><input type="hidden" name="member_id" value="{{m.id}}"/><button class="btn br bsm">✕ Reject</button></form>
  </td>
</tr>{% endfor %}</tbody>
</table></div>{% endif %}
</div></div>

<!-- PAYMENTS -->
<div id="tab-pay" class="tp"><div class="card"><h2>Payment Ledger</h2>
<h3>Add / Update Payment</h3>
<form method="post" action="/admin/payment/update" style="display:grid;grid-template-columns:repeat(auto-fit,minmax(140px,1fr));gap:12px;margin-bottom:24px;">
  <div class="fg" style="margin:0;"><label>Member ID</label><input name="member_id" placeholder="ID"/></div>
  <div class="fg" style="margin:0;"><label>Month &amp; Year</label>
    <select name="amo" id="asel" onchange="sa(this)">
      {% for m,y in month_options %}<option value="{{m}}|{{y}}">{{mfull[m]}} {{y}}</option>{% endfor %}
    </select>
    <input type="hidden" name="month" id="amo_m" value="{{month_options[0][0] if month_options else ''}}"/>
    <input type="hidden" name="year" id="amo_y" value="{{month_options[0][1] if month_options else ''}}"/>
    <script>function sa(s){var p=s.value.split('|');document.getElementById('amo_m').value=p[0];document.getElementById('amo_y').value=p[1];}
    var as2=document.getElementById('asel');if(as2)sa(as2);</script>
  </div>
  <div class="fg" style="margin:0;"><label>Status</label>
    <select name="status"><option value="Verified">Verified</option><option value="Pending">Pending</option><option value="Rejected">Rejected</option></select>
  </div>
  <div class="fg" style="margin:0;"><label>TxnID</label><input name="txn_id" placeholder="Transaction ID"/></div>
  <div style="display:flex;align-items:flex-end;"><button class="btn bg" type="submit">Save</button></div>
</form>
<hr class="fn"/>
<h3>Pending Verifications ({{pending_p|length}})</h3>
{% if pending_p %}<div class="tw"><table>
<thead><tr><th>Member</th><th>Month</th><th>Year</th><th>TxnID</th><th>Submitted</th><th>Action</th></tr></thead>
<tbody>{% for p in pending_p %}<tr>
  <td>{{p.member_id}}</td><td>{{mfull.get(p.month|int,p.month)}}</td><td>{{p.year}}</td>
  <td><code>{{p.txn_id}}</code></td><td style="font-size:.8rem;">{{p.submitted_at}}</td>
  <td style="display:flex;gap:4px;">
    <form method="post" action="/admin/payment/verify">
      <input type="hidden" name="pay_id" value="{{p.id}}"/>
      <button class="btn bgrn bsm">✓ Verify</button></form>
    <form method="post" action="/admin/payment/reject_p">
      <input type="hidden" name="pay_id" value="{{p.id}}"/>
      <button class="btn br bsm">✕</button></form>
  </td>
</tr>{% endfor %}</tbody>
</table></div>{% else %}<p style="color:var(--sl);">No pending submissions.</p>{% endif %}
</div></div>

<!-- ALL MEMBERS -->
<div id="tab-mem" class="tp"><div class="card"><h2>All Members ({{all_members|length}})</h2>
<div class="tw"><table>
<thead><tr><th>ID</th><th>Name</th><th>Phone</th><th>Email</th><th>Dept</th><th>Blood</th><th>DOB</th><th>Status</th><th>Joined</th><th>Actions</th></tr></thead>
<tbody>{% for m in all_members %}<tr>
  <td><code>{{m.id}}</code></td><td>{{m.name}}</td><td>{{m.phone}}</td><td>{{m.email}}</td>
  <td>{{m.department or '—'}}</td><td>{{m.blood_group or '—'}}</td><td>{{m.dob or '—'}}</td>
  <td>{% if m.status=='Approved' %}<span class="bdg bdgg">Approved</span>
      {% elif m.status=='Pending' %}<span class="bdg bdgy">Pending</span>
      {% else %}<span class="bdg bdgr">{{m.status}}</span>{% endif %}</td>
  <td>{{m.join_date}}</td>
  <td style="display:flex;gap:4px;flex-wrap:wrap;">
    <a href="/admin/member/{{m.id}}/print" target="_blank" class="btn bbl bsm">🖨</a>
    <form method="post" action="/admin/member/remove" onsubmit="return confirm('Remove {{m.name}}?');">
      <input type="hidden" name="member_id" value="{{m.id}}"/>
      <button class="btn br bsm">✕</button>
    </form>
  </td>
</tr>{% endfor %}</tbody>
</table></div>
</div></div>

<!-- PAYMENT SUMMARY -->
<div id="tab-sum" class="tp">
<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:16px;margin-bottom:24px;">
  <div class="sc" style="background:linear-gradient(135deg,#1a3a55,#2c5f8a);color:#fff;"><div class="v">{{total_col}} ৳</div><div class="l" style="color:rgba(255,255,255,.7);">Collected</div></div>
  <div class="sc" style="background:linear-gradient(135deg,#155724,#27ae60);color:#fff;"><div class="v">{{total_ver}}</div><div class="l" style="color:rgba(255,255,255,.7);">Verified</div></div>
  <div class="sc" style="background:linear-gradient(135deg,#7d6608,#b8860b);color:#fff;"><div class="v">{{total_pend_amt}} ৳</div><div class="l" style="color:rgba(255,255,255,.7);">Pending</div></div>
  <div class="sc" style="background:linear-gradient(135deg,#721c24,#c0392b);color:#fff;"><div class="v">{{total_due_amt}} ৳</div><div class="l" style="color:rgba(255,255,255,.7);">Outstanding</div></div>
  <div class="sc" style="background:linear-gradient(135deg,#333,#555);color:#fff;"><div class="v">{{total_exp|int}} ৳</div><div class="l" style="color:rgba(255,255,255,.7);">Expenses</div></div>
  <div class="sc" style="background:linear-gradient(135deg,#145a32,#1e8449);color:#fff;"><div class="v">{{cur_bal|int}} ৳</div><div class="l" style="color:rgba(255,255,255,.7);">Balance</div></div>
</div>
<div class="card"><h2>Add Expense</h2>
<form method="post" action="/admin/expenses/add" style="display:grid;grid-template-columns:2fr 1fr 1fr 1fr auto;gap:12px;align-items:flex-end;">
  <div class="fg" style="margin:0;"><label>Title</label><input name="title" required placeholder="Expense title"/></div>
  <div class="fg" style="margin:0;"><label>Amount (৳)</label><input name="amount" type="number" step="0.01" required/></div>
  <div class="fg" style="margin:0;"><label>Date</label><input name="date" type="date" required value="{{now.strftime('%Y-%m-%d')}}"/></div>
  <div class="fg" style="margin:0;"><label>Notes</label><input name="notes" placeholder="Optional"/></div>
  <div style="padding-bottom:2px;"><button class="btn bg" type="submit">Add</button></div>
</form>
{% if expenses %}<hr class="fn"/><h3>Expense Log</h3>
<div class="tw"><table>
<thead><tr><th>#</th><th>Title</th><th>Amount</th><th>Date</th><th>Notes</th><th>Del</th></tr></thead>
<tbody>{% for e in expenses %}<tr>
  <td>{{loop.index}}</td><td>{{e.title}}</td><td style="color:var(--red);font-weight:600;">{{e.amount}} ৳</td>
  <td>{{e.date}}</td><td>{{e.notes or '—'}}</td>
  <td><form method="post" action="/admin/expenses/delete"><input type="hidden" name="expense_id" value="{{e.id}}"/><button class="btn br bsm">✕</button></form></td>
</tr>{% endfor %}</tbody>
</table></div>{% endif %}
</div>
<div class="card"><h2>📧 Send Payment Due Notices</h2>
<div class="al ali">Uses the Gmail credentials saved in the <strong>Email</strong> tab. Individual 📧 buttons below use the same saved credentials.</div>
<form method="post" action="/admin/notices/send_all" id="nform">
  <button class="btn bg" type="submit">📧 Send All Due Notices</button>
</form>
</div>
<div class="card"><h2>Per-Member Breakdown</h2>
<div class="tw"><table>
<thead><tr><th>ID</th><th>Name</th>
{% for lbl in summary_months %}<th style="font-size:.68rem;white-space:nowrap;">{{lbl}}</th>{% endfor %}
<th>Paid</th><th>Due</th><th>৳</th><th>Notify</th></tr></thead>
<tbody>{% for row in member_summary %}<tr>
  <td><code>{{row.id}}</code></td><td>{{row.name}}</td>
  {% for cell in row.cells %}
    {% if cell=='V' %}<td style="text-align:center;color:#27ae60;font-weight:700;">✓</td>
    {% elif cell=='P' %}<td style="text-align:center;color:#b8860b;font-weight:700;">⏳</td>
    {% else %}<td style="text-align:center;color:#c0392b;font-weight:700;">✗</td>{% endif %}
  {% endfor %}
  <td style="color:#27ae60;font-weight:700;">{{row.paid_count}}</td>
  <td style="color:#c0392b;font-weight:700;">{{row.due_count}}</td>
  <td style="font-weight:700;">{{row.paid_amount}} ৳</td>
  <td>{% if row.due_count>0 and row.email %}
    <form method="post" action="/admin/notices/send_one">
      <input type="hidden" name="member_id" value="{{row.id}}"/>
      <button class="btn bsm" style="background:#e67e22;color:#fff;" type="submit">📧</button>
    </form>{% else %}—{% endif %}</td>
</tr>{% endfor %}</tbody>
</table></div>
<p style="margin-top:10px;font-size:.78rem;color:var(--sl);">✓ Verified &nbsp;|&nbsp; ⏳ Pending &nbsp;|&nbsp; ✗ Unpaid</p>
</div></div>

<!-- DB EDITOR -->
<div id="tab-db" class="tp">
<div class="card" style="background:linear-gradient(135deg,#1a3a55,#2c3e50);color:#fff;">
  <h2 style="color:#fff;border-color:#444;">📊 Excel Database Tools</h2>
  <div style="display:flex;gap:10px;margin-bottom:18px;flex-wrap:wrap;">
    <a href="/admin/export/xlsx" class="btn bgrn">📥 Export Full Database (.xlsx)</a>
  </div>
  <hr style="border:none;border-top:1px solid rgba(255,255,255,.2);margin:18px 0;"/>
  <h3 style="color:#fff;">📤 Import from Excel</h3>
  <p style="color:rgba(255,255,255,.75);font-size:.9rem;margin-bottom:14px;">
    Upload a <strong>.xlsx</strong> file to bulk-import data into Firestore. Supported sheets:<br/><br/>
    <strong>Members</strong> sheet columns:<br/>
    <code style="font-size:.8rem;">ID | Name | Phone | Email | Password | University | Department | BloodGroup | DateOfBirth | Address | Status | JoinDate</code><br/><br/>
    <strong>Payments</strong> sheet columns:<br/>
    <code style="font-size:.8rem;">MemberID | Month | Year | Status | TxnID | SubmittedAt | VerifiedAt</code><br/><br/>
    <span style="font-size:.8rem;color:rgba(255,255,255,.5);">Existing records with same ID will be updated. Password = plain text, auto-hashed.</span>
  </p>
  <form method="post" action="/admin/import/xlsx" enctype="multipart/form-data">
    <div style="display:flex;gap:10px;align-items:center;flex-wrap:wrap;">
      <input type="file" name="xlsx_file" accept=".xlsx" required
             style="flex:1;padding:8px 12px;border-radius:6px;border:none;font-size:.85rem;background:rgba(255,255,255,.15);color:#fff;min-width:0;"/>
      <button class="btn bg" type="submit"
              onclick="return confirm('This will import/update data from your Excel file. Continue?');">
        📤 Import Now
      </button>
    </div>
  </form>
</div>
<div class="card"><h2>✏️ Live Member Editor</h2>
<div class="al ali">Search a member by ID to edit their details directly.</div>
<div style="display:flex;gap:10px;margin-bottom:20px;">
  <input id="edit_mid" style="flex:1;padding:10px 14px;border:1.5px solid var(--bdr);border-radius:6px;" placeholder="Enter Member ID to edit…"/>
  <button class="btn bg" onclick="loadMember()">Load</button>
</div>
<div id="member_edit_area"><p style="color:var(--sl);">Enter a Member ID above to edit their record.</p></div>
</div>
<div class="card"><h2>✏️ Live Payment Editor</h2>
<div class="al ali">Search by Member ID to view and edit their payment records.</div>
<div style="display:flex;gap:10px;margin-bottom:20px;">
  <input id="edit_pmid" style="flex:1;padding:10px 14px;border:1.5px solid var(--bdr);border-radius:6px;" placeholder="Enter Member ID…"/>
  <button class="btn bg" onclick="loadPayments()">Load</button>
</div>
<div id="payment_edit_area"><p style="color:var(--sl);">Enter a Member ID above to view payments.</p></div>
</div>
<script>
function loadMember(){
  var mid=document.getElementById('edit_mid').value.trim().toLowerCase();
  if(!mid)return;
  fetch('/admin/db/member/'+mid).then(r=>r.json()).then(function(d){
    if(d.error){document.getElementById('member_edit_area').innerHTML='<div class="al ale">'+d.error+'</div>';return;}
    var m=d.member;
    var html='<form method="post" action="/admin/db/member/save">';
    html+='<input type="hidden" name="original_id" value="'+m.id+'"/>';
    var fields=[['id','Member ID'],['name','Full Name'],['phone','Phone'],['email','Email'],
                ['university','University'],['department','Department'],['blood_group','Blood Group'],
                ['dob','Date of Birth'],['address','Address'],['status','Status (Approved/Pending/Rejected)'],['join_date','Join Date']];
    html+='<div class="g2">';
    if(m.photo_url){
      html+='<div class="fg" style="grid-column:1/-1;text-align:center;"><img src="'+m.photo_url+'" style="width:80px;height:80px;border-radius:50%;object-fit:cover;border:3px solid #C9A84C;"/></div>';
    }
    fields.forEach(function(f){
      html+='<div class="fg"><label>'+f[1]+'</label><input name="'+f[0]+'" value="'+(m[f[0]]||'').toString().replace(/"/g,"&quot;")+'"/></div>';
    });
    html+='</div>';
    html+='<div class="fg"><label>New Password (leave blank to keep)</label><input type="password" name="new_password" placeholder="Leave blank to keep current password"/></div>';
    html+='<button class="btn bg" type="submit">💾 Save Member</button></form>';
    document.getElementById('member_edit_area').innerHTML=html;
  }).catch(function(e){document.getElementById('member_edit_area').innerHTML='<div class="al ale">Error: '+e+'</div>';});
}
function loadPayments(){
  var mid=document.getElementById('edit_pmid').value.trim().toLowerCase();
  if(!mid)return;
  fetch('/admin/db/payments/'+mid).then(r=>r.json()).then(function(d){
    if(d.error){document.getElementById('payment_edit_area').innerHTML='<div class="al ale">'+d.error+'</div>';return;}
    var ps=d.payments;
    if(!ps.length){document.getElementById('payment_edit_area').innerHTML='<p style="color:var(--sl);">No payment records found for this member.</p>';return;}
    var html='<div class="tw"><table><thead><tr><th>Month</th><th>Year</th><th>Status</th><th>TxnID</th><th>Submitted</th><th>Action</th></tr></thead><tbody>';
    ps.forEach(function(p){
      html+='<tr><td>'+p.month+'</td><td>'+p.year+'</td>';
      html+='<td><form method="post" action="/admin/db/payment/update" style="display:flex;gap:6px;">';
      html+='<input type="hidden" name="pay_id" value="'+p.id+'"/>';
      html+='<select name="status" style="padding:4px;border-radius:4px;border:1px solid #ddd;font-size:.82rem;">';
      ['Verified','Pending','Rejected'].forEach(function(s){html+='<option'+(p.status==s?' selected':'')+'>'+s+'</option>';});
      html+='</select>';
      html+='<input name="txn_id" value="'+(p.txn_id||'')+'" style="width:130px;padding:4px 6px;border:1px solid #ddd;border-radius:4px;font-size:.82rem;" placeholder="TxnID"/>';
      html+='<button class="btn bgrn bsm" type="submit">💾</button></form></td>';
      html+='<td><code>'+(p.txn_id||'—')+'</code></td><td style="font-size:.8rem;">'+(p.submitted_at||'—')+'</td>';
      html+='<td><form method="post" action="/admin/db/payment/delete"><input type="hidden" name="pay_id" value="'+p.id+'"/><button class="btn br bsm" onclick="return confirm(\'Delete this record?\')">✕</button></form></td>';
      html+='</tr>';
    });
    html+='</tbody></table></div>';
    document.getElementById('payment_edit_area').innerHTML=html;
  }).catch(function(e){document.getElementById('payment_edit_area').innerHTML='<div class="al ale">Error: '+e+'</div>';});
}
</script>
</div>

<!-- HOME CONTENT -->
<div id="tab-cnt" class="tp"><div class="card"><h2>Edit Home Content</h2>
<form method="post" action="/admin/content/update">
  <div class="fg"><label>About the Club</label><textarea name="about" rows="4">{{content.get('about','')}}</textarea></div>
  <div class="fg"><label>Announcements</label><textarea name="announcements" rows="3">{{content.get('announcements','')}}</textarea></div>
  <div class="fg"><label>Notice Board</label><textarea name="notice" rows="3">{{content.get('notice','')}}</textarea></div>
  <div class="fg" style="display:flex;align-items:center;gap:8px;">
    <input type="checkbox" name="notify_members" id="notify_members" value="1" style="width:auto;" checked/>
    <label for="notify_members" style="margin:0;text-transform:none;font-weight:500;color:var(--ink);">📧 Email this notice to all approved members</label>
  </div>
  <button class="btn bg" type="submit">Update Content</button>
</form></div></div>

<!-- GALLERY -->
<div id="tab-gal" class="tp"><div class="card"><h2>Gallery Management</h2>
<form method="post" action="/admin/gallery/upload" enctype="multipart/form-data">
  <div style="display:flex;gap:12px;align-items:flex-end;flex-wrap:wrap;">
    <div class="fg" style="flex:1;margin:0;"><label>Image File</label><input type="file" name="image" accept="image/*" required/></div>
    <div class="fg" style="flex:1;margin:0;"><label>Caption</label><input name="caption" placeholder="Optional"/></div>
    <button class="btn bg" type="submit">Upload</button>
  </div>
</form>
<hr class="fn"/>
{% if gallery %}<div class="gg">{% for img in gallery %}
  <div style="position:relative;border-radius:10px;overflow:hidden;aspect-ratio:1;">
    <img src="{{img.url}}" style="width:100%;height:100%;object-fit:cover;cursor:pointer;" onclick="openLB('{{img.url}}')" alt="{{img.caption}}"/>
    <form method="post" action="/admin/gallery/delete" style="position:absolute;top:4px;right:4px;">
      <input type="hidden" name="image_id" value="{{img.id}}"/>
      <button class="btn br bsm" style="padding:3px 8px;">✕</button>
    </form>
  </div>{% endfor %}
</div>{% else %}<p style="color:var(--sl);">No images yet.</p>{% endif %}
</div></div>

<!-- EVENTS -->
<div id="tab-ev" class="tp"><div class="card"><h2>Manage Events</h2>
<form method="post" action="/admin/events/add" enctype="multipart/form-data">
  <div class="g2">
    <div class="fg"><label>Title</label><input name="title" required/></div>
    <div class="fg"><label>Date</label><input name="event_date" type="date" required/></div>
  </div>
  <div class="fg"><label>Description</label><textarea name="description" rows="3"></textarea></div>
  <div class="fg"><label>Banner Image</label><input type="file" name="image" accept="image/*"/></div>
  <button class="btn bg" type="submit">Add Event</button>
</form>
<hr class="fn"/>
{% if events %}<div class="tw"><table>
<thead><tr><th>Title</th><th>Date</th><th>Description</th><th>Del</th></tr></thead>
<tbody>{% for ev in events %}<tr>
  <td>{{ev.title}}</td><td>{{ev.event_date}}</td>
  <td style="max-width:260px;font-size:.85rem;">{{(ev.description or '')[:80]}}{% if ev.description and ev.description|length>80 %}…{% endif %}</td>
  <td><form method="post" action="/admin/events/delete"><input type="hidden" name="event_id" value="{{ev.id}}"/><button class="btn br bsm">Del</button></form></td>
</tr>{% endfor %}</tbody>
</table></div>{% endif %}
</div></div>

<!-- BROADCAST EMAIL -->
<div id="tab-eml" class="tp">
<div class="card" style="max-width:600px;width:100%;background:linear-gradient(135deg,#1a3a55,#2c3e50);color:#fff;">
  <h2 style="color:#fff;border-color:#444;">💾 Saved Gmail Credentials</h2>
  <p style="color:rgba(255,255,255,.75);font-size:.85rem;margin-bottom:14px;">
    Enter your Gmail address and App Password once — they'll be remembered for Broadcast Email, Due Notices, and pre-filled everywhere until you change them here.
  </p>
  <form method="post" action="/admin/email/save_creds">
    <div class="g2">
      <div class="fg" style="margin-bottom:8px;"><label style="color:rgba(255,255,255,.7);">Gmail Address</label>
        <input name="sender_email" type="email" required value="{{saved_creds.sender_email or club_email}}" style="background:rgba(255,255,255,.92);"/></div>
      <div class="fg" style="margin-bottom:8px;"><label style="color:rgba(255,255,255,.7);">Gmail App Password</label>
        <input name="sender_password" type="password" required placeholder="{{ '••••••••••••••••' if saved_creds.sender_password else '16-char App Password' }}" style="background:rgba(255,255,255,.92);"/></div>
    </div>
    <button class="btn bg" type="submit">💾 Save Credentials</button>
  </form>
</div>

<div class="card" style="max-width:600px;width:100%;background:linear-gradient(135deg,#7d6608,#b8860b);color:#fff;">
  <h2 style="color:#fff;border-color:rgba(255,255,255,.3);">🔔 Admin Notification Email</h2>
  <p style="color:rgba(255,255,255,.8);font-size:.85rem;margin-bottom:14px;">
    Get notified automatically whenever a member submits a payment or a new registration form is submitted.
  </p>
  <form method="post" action="/admin/email/save_notify">
    <div class="fg" style="margin-bottom:8px;"><label style="color:rgba(255,255,255,.8);">Notify This Email</label>
      <input name="notify_email" type="email" required value="{{notify_email}}" placeholder="your-personal@gmail.com" style="background:rgba(255,255,255,.92);"/></div>
    <button class="btn bp" type="submit">Save Notification Email</button>
  </form>
  <p style="color:rgba(255,255,255,.6);font-size:.78rem;margin-top:10px;">Uses the saved Gmail credentials above to send the notification.</p>
</div>

<div class="card" style="max-width:600px;width:100%;"><h2>Broadcast Email</h2>
<div class="al ali">Sends to all approved members using your saved Gmail credentials above.</div>
<form method="post" action="/admin/email/broadcast">
  <div class="fg"><label>Subject</label><input name="subject" required/></div>
  <div class="fg"><label>Message</label><textarea name="body" rows="6" required></textarea></div>
  <button class="btn bg" type="submit">Send to All Members</button>
</form></div></div>

<!-- LOGO & SETTINGS -->
<div id="tab-logo" class="tp">
<div class="card" style="max-width:480px;width:100%;">
  <h2>Club Logo</h2>
  {% if nav_logo %}
  <div style="text-align:center;margin-bottom:20px;padding:20px;background:var(--ink);border-radius:10px;">
    <img src="{{nav_logo}}" style="max-height:140px;max-width:100%;object-fit:contain;filter:brightness(0) invert(1);" alt="Logo"/>
    <p style="margin-top:8px;font-size:.8rem;color:#aaa;">Preview — as shown in navbar</p>
  </div>{% else %}<div class="al ali">No logo uploaded yet.</div>{% endif %}
  <form method="post" action="/admin/logo/upload" enctype="multipart/form-data">
    <div class="fg"><label>Upload New Logo</label><input type="file" name="logo" accept="image/*" required/></div>
    <p style="font-size:.8rem;color:var(--sl);margin-bottom:12px;">Black logo will appear white on the dark navbar automatically. This logo is also used as the site favicon.</p>
    <button class="btn bg" type="submit">Upload Logo</button>
  </form>
</div>
<div class="card" style="max-width:480px;width:100%;">
  <h2>🖼️ Website Background Image</h2>
  {% if site_bg %}
  <div style="margin-bottom:16px;border-radius:10px;overflow:hidden;">
    <img src="{{site_bg}}" style="width:100%;height:120px;object-fit:cover;" alt="Background"/>
  </div>{% else %}<div class="al ali">No custom background set — using default cream color.</div>{% endif %}
  <form method="post" action="/admin/settings/bg_upload" enctype="multipart/form-data">
    <div class="fg"><label>Upload Background Image</label><input type="file" name="bg_image" accept="image/*" required/></div>
    <button class="btn bg" type="submit">Upload Background</button>
  </form>
  {% if site_bg %}
  <form method="post" action="/admin/settings/bg_remove" style="margin-top:8px;">
    <button class="btn br bsm" type="submit">Remove Background</button>
  </form>{% endif %}
</div>
<div class="card" style="max-width:480px;width:100%;">
  <h2>📱 Mobile View Scale</h2>
  <p style="font-size:.85rem;color:var(--sl);margin-bottom:14px;">Adjust the initial zoom level for mobile devices only. Desktop view is not affected. Default is 1.0.</p>
  <form method="post" action="/admin/settings/mobile_scale">
    <div class="fg"><label>Mobile Initial Scale</label>
      <input name="mobile_scale" type="number" step="0.05" min="0.5" max="2.0" value="{{mobile_scale}}" required/>
    </div>
    <button class="btn bg" type="submit">Save Scale</button>
  </form>
</div>
<div class="card" style="max-width:480px;width:100%;">
  <h2>🔐 Change Admin Password</h2>
  <div class="al ali">You can always recover admin access if you forget your password. Contact your system administrator.</div>
  <form method="post" action="/admin/change_password">
    <div class="fg"><label>Current Password</label><input name="current_password" type="password" required/></div>
    <div class="fg"><label>New Password</label><input name="new_password" type="password" required/></div>
    <div class="fg"><label>Confirm New Password</label><input name="confirm_password" type="password" required/></div>
    <button class="btn bp" type="submit">Change Password</button>
  </form>
</div>
</div>
"""
    site = get_site_settings()
    return render(tmpl, _title="Admin Dashboard — Civision Society",
                  pending=pending, approved=approved, all_members=members,
                  pending_p=pending_p, events=events, gallery=gallery,
                  content=content, month_options=month_options, mfull=MONTHS_FULL,
                  nav_logo=nav_logo, summary_months=summary_months,
                  member_summary=member_summary, total_col=total_col,
                  site_bg=site.get("bg_image",""), mobile_scale=site.get("mobile_scale","1.0"),
                  total_ver=total_ver, total_pend_amt=total_pend_amt,
                  total_due_amt=total_due_amt, expenses=expenses,
                  total_exp=total_exp, cur_bal=cur_bal,
                  saved_creds=get_saved_email_creds(), notify_email=get_notify_email())

# ── ADMIN ACTIONS ─────────────────────────────────────────────────────────────
@app.route("/admin/export/xlsx")
@admin_required
def admin_export_xlsx():
    import io
    try:
        from openpyxl import Workbook
    except ImportError:
        return "openpyxl not installed. Add it to requirements.txt", 500

    wb = Workbook()

    # Members sheet
    ws_m = wb.active; ws_m.title = "Members"
    ws_m.append(["ID","Name","Phone","Email","University","Department",
                 "BloodGroup","DateOfBirth","Address","Status","JoinDate"])
    for m in get_all_members():
        ws_m.append([m.get("id",""), m.get("name",""), m.get("phone",""),
                     m.get("email",""), m.get("university",""), m.get("department",""),
                     m.get("blood_group",""), m.get("dob",""), m.get("address",""),
                     m.get("status",""), m.get("join_date","")])

    # Payments sheet
    ws_p = wb.create_sheet("Payments")
    ws_p.append(["PayID","MemberID","Month","Year","Status","TxnID","SubmittedAt","VerifiedAt"])
    for p in get_payments():
        ws_p.append([p.get("id",""), p.get("member_id",""), p.get("month",""),
                     p.get("year",""), p.get("status",""), p.get("txn_id",""),
                     p.get("submitted_at",""), p.get("verified_at","")])

    # Expenses sheet
    ws_e = wb.create_sheet("Expenses")
    ws_e.append(["ID","Title","Amount","Date","Notes"])
    for e in get_expenses():
        ws_e.append([e.get("id",""), e.get("title",""), e.get("amount",""),
                     e.get("date",""), e.get("notes","")])

    # Events sheet
    ws_ev = wb.create_sheet("Events")
    ws_ev.append(["ID","Title","Description","EventDate","ImageURL","CreatedAt"])
    for ev in get_events():
        ws_ev.append([ev.get("id",""), ev.get("title",""), ev.get("description",""),
                      ev.get("event_date",""), ev.get("image_url",""), ev.get("created_at","")])

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"CivisionDB_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/admin/import/xlsx", methods=["POST"])
@admin_required
def admin_import_xlsx():
    if "xlsx_file" not in request.files:
        flash("No file selected.","e"); return redirect(url_for("admin_dashboard")+"#tab-db")
    f = request.files["xlsx_file"]
    if not f or not f.filename.endswith(".xlsx"):
        flash("Only .xlsx files accepted.","e"); return redirect(url_for("admin_dashboard")+"#tab-db")
    try:
        from openpyxl import load_workbook
        import io
        buf = io.BytesIO(f.read())
        wb = load_workbook(buf)
    except Exception as e:
        flash(f"Could not read Excel file: {str(e)}","e"); return redirect(url_for("admin_dashboard")+"#tab-db")

    imported = updated = errors = 0

    if "Members" in wb.sheetnames:
        ws = wb["Members"]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            hdr = [str(h).strip() if h else "" for h in rows[0]]
            def g(row, name):
                for variant in [name, name.lower(), name.upper(),
                                 name.replace(" ",""), name.replace("_","")]:
                    if variant in hdr:
                        i = hdr.index(variant)
                        return str(row[i]).strip() if i < len(row) and row[i] is not None else ""
                return ""
            for row in rows[1:]:
                if not row[0]: continue
                try:
                    mid = str(row[0]).strip().lower()
                    if not mid: continue
                    pw_plain = g(row,"Password") or g(row,"password") or g(row,"PasswordHash")
                    if pw_plain and len(pw_plain)==64 and all(c in "0123456789abcdef" for c in pw_plain.lower()):
                        pw_hash = pw_plain
                    elif pw_plain:
                        pw_hash = hash_pw(pw_plain)
                    else:
                        existing = get_member_by_id(mid)
                        pw_hash = existing.get("password_hash","") if existing else hash_pw("changeme")
                    existing = get_member_by_id(mid)
                    data = {
                        "id":           mid,
                        "name":         g(row,"Name") or g(row,"name"),
                        "phone":        g(row,"Phone") or g(row,"phone"),
                        "email":        g(row,"Email") or g(row,"email"),
                        "password_hash":pw_hash,
                        "university":   g(row,"University") or g(row,"university"),
                        "department":   g(row,"Department") or g(row,"department"),
                        "blood_group":  g(row,"BloodGroup") or g(row,"blood_group") or g(row,"Blood Group"),
                        "dob":          g(row,"DateOfBirth") or g(row,"dob") or g(row,"Date of Birth"),
                        "address":      g(row,"Address") or g(row,"address"),
                        "status":       g(row,"Status") or g(row,"status") or "Pending",
                        "join_date":    g(row,"JoinDate") or g(row,"join_date") or g(row,"Join Date") or datetime.now().strftime("%Y-%m-%d"),
                    }
                    if existing:
                        data["photo_url"]     = existing.get("photo_url","")
                        data["session_token"] = existing.get("session_token","")
                    save_member(mid, data)
                    if existing: updated += 1
                    else: imported += 1
                except Exception:
                    errors += 1

    # Import Payments sheet
    pay_imported = pay_updated = pay_errors = 0
    if "Payments" in wb.sheetnames:
        ws_p = wb["Payments"]
        p_rows = list(ws_p.iter_rows(values_only=True))
        if p_rows:
            p_hdr = [str(h).strip() if h else "" for h in p_rows[0]]
            def gp(row, *names):
                for name in names:
                    for variant in [name, name.lower(), name.replace(" ",""), name.replace("_","")]:
                        if variant in p_hdr:
                            i = p_hdr.index(variant)
                            return str(row[i]).strip() if i < len(row) and row[i] is not None else ""
                return ""
            for row in p_rows[1:]:
                if not any(row): continue
                try:
                    mid   = gp(row,"MemberID","member_id","Member ID","ID").lower()
                    month = gp(row,"Month","month")
                    year  = gp(row,"Year","year")
                    if not mid or not month or not year: pay_errors+=1; continue
                    status  = gp(row,"Status","status") or "Verified"
                    txn_id  = gp(row,"TxnID","txn_id","Transaction ID","TransactionID")
                    sub_at  = gp(row,"SubmittedAt","submitted_at","Submitted At") or datetime.now().strftime("%Y-%m-%d %H:%M")
                    ver_at  = gp(row,"VerifiedAt","verified_at","Verified At") or (datetime.now().strftime("%Y-%m-%d %H:%M") if status=="Verified" else "")
                    pid = f"{mid}_{month}_{year}"
                    existing_p = col("payments").document(pid).get()
                    save_payment(pid, {
                        "member_id":   mid,
                        "month":       int(float(month)),
                        "year":        int(float(year)),
                        "status":      status,
                        "txn_id":      txn_id,
                        "submitted_at":sub_at,
                        "verified_at": ver_at,
                    })
                    if existing_p.exists: pay_updated+=1
                    else: pay_imported+=1
                except Exception:
                    pay_errors+=1

    msg = f"Members: {imported} added, {updated} updated"
    if errors: msg += f", {errors} skipped"
    msg += f" | Payments: {pay_imported} added, {pay_updated} updated"
    if pay_errors: msg += f", {pay_errors} skipped"
    flash(msg, "s")
    return redirect(url_for("admin_dashboard")+"#tab-db")


@app.route("/admin/approve", methods=["POST"])
@admin_required
def admin_approve():
    mid = request.form.get("member_id","").strip().lower()
    update_member(mid, {"status":"Approved"})
    flash(f"Member {mid} approved.","s")
    return redirect(url_for("admin_dashboard")+"#tab-apr")

@app.route("/admin/reject", methods=["POST"])
@admin_required
def admin_reject():
    mid = request.form.get("member_id","").strip().lower()
    update_member(mid, {"status":"Rejected"})
    flash(f"Member {mid} rejected.","i")
    return redirect(url_for("admin_dashboard")+"#tab-apr")

@app.route("/admin/member/remove", methods=["POST"])
@admin_required
def admin_remove():
    mid = request.form.get("member_id","").strip().lower()
    col("members").document(mid).delete()
    for p in get_payments_for_member(mid):
        delete_payment(p["id"])
    flash(f"Member {mid} removed.","s")
    return redirect(url_for("admin_dashboard")+"#tab-mem")

@app.route("/admin/member/<mid>/print")
@admin_required
def admin_print(mid):
    member = get_member_by_id(mid.lower())
    if not member: return "Not found.", 404
    ps  = get_payment_status_for_member(mid.lower())
    ver = [p for p in ps if p["status"]=="Verified"]
    ped = [p for p in ps if p["status"]=="Pending"]
    unp = [p for p in ps if p["status"]=="Unpaid"]
    tp  = len(ver)*MONTHLY_DUES; td = len(unp)*MONTHLY_DUES
    logo = get_nav_logo()
    def _rc(s): return '#1a5c38' if s=='Verified' else '#7d6608' if s=='Pending' else '#c0392b'
    def _rl(s): return '&#10003; Verified' if s=='Verified' else '&#9203; Pending' if s=='Pending' else '&#10007; Unpaid'
    rows = ""
    for i,p in enumerate(ps):
        st = p["status"]; c = _rc(st); l = _rl(st); txn = p["txn_id"] or "—"
        rows += f"<tr><td>{i+1}</td><td>{p['month']}</td><td>150 &#2547;</td><td style=\"color:{c};font-weight:700;\">{l}</td><td><code>{txn}</code></td></tr>"
    bal = (f'<div class="warn">Outstanding: <strong>{td} ৳</strong> ({len(unp)} months)</div>'
           if unp else '<div class="ok">✓ All dues cleared.</div>')
    return f"""<!DOCTYPE html><html><head><meta charset="UTF-8"/>
<title>Report — {member.get('name')}</title>
<link href="https://fonts.googleapis.com/css2?family=Cormorant+Garamond:wght@700&family=Inter:wght@400;600&display=swap" rel="stylesheet"/>
<style>
*{{box-sizing:border-box;margin:0;padding:0;}}
body{{font-family:'Inter',sans-serif;padding:32px;font-size:13px;color:#1a1a1a;}}
.hdr{{display:flex;align-items:center;gap:20px;border-bottom:3px solid #1a1a1a;padding-bottom:20px;margin-bottom:24px;}}
.hdr img.logo{{height:70px;width:70px;object-fit:contain;}}
.hdr img.photo{{height:80px;width:80px;object-fit:cover;border-radius:50%;border:3px solid #C9A84C;margin-left:auto;}}
.hdr .photo-ph{{height:80px;width:80px;border-radius:50%;border:3px solid #C9A84C;margin-left:auto;background:#1a1a1a;color:#C9A84C;
                display:flex;align-items:center;justify-content:center;font-family:'Cormorant Garamond',serif;font-size:2rem;font-weight:700;}}
.hdr h1{{font-family:'Cormorant Garamond',serif;font-size:2rem;}}
.hdr p{{color:#777;font-size:.82rem;text-transform:uppercase;letter-spacing:.08em;}}
.sec{{font-family:'Cormorant Garamond',serif;font-size:1.2rem;font-weight:700;margin:20px 0 10px;border-bottom:1px solid #ddd;padding-bottom:6px;}}
.ig{{display:grid;grid-template-columns:1fr 1fr;gap:10px 24px;margin-bottom:20px;}}
.ig label{{font-size:.7rem;text-transform:uppercase;letter-spacing:.08em;color:#888;display:block;}}
.ig span{{font-size:.95rem;font-weight:600;}}
.boxes{{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-bottom:20px;}}
.box{{border:2px solid #ddd;border-radius:8px;padding:14px;text-align:center;}}
.box .bv{{font-size:1.5rem;font-weight:700;}} .box .bl{{font-size:.68rem;text-transform:uppercase;letter-spacing:.06em;color:#888;margin-top:4px;}}
.ok{{background:#d5f5e3;border:1px solid #a9dfbf;padding:10px 14px;border-radius:6px;margin-bottom:18px;color:#1a5c38;font-size:.85rem;}}
.warn{{background:#fef9e7;border:1px solid #f9e79f;padding:10px 14px;border-radius:6px;margin-bottom:18px;color:#7d6608;font-size:.85rem;}}
table{{width:100%;border-collapse:collapse;font-size:.85rem;}}
th{{background:#1a1a1a;color:#fff;padding:8px 12px;text-align:left;font-size:.75rem;text-transform:uppercase;}}
td{{padding:7px 12px;border-bottom:1px solid #eee;}}
.footer{{border-top:1px solid #ddd;padding-top:12px;text-align:center;color:#aaa;font-size:.75rem;margin-top:20px;}}
.np{{margin-top:20px;text-align:center;}}
@media print{{.np{{display:none;}}body{{padding:16px;}}}}
</style></head><body>
<div class="hdr">
  {'<img class="logo" src="'+logo+'" alt="Logo"/>' if logo else ''}
  <div><h1>Civision Society</h1><p>"Connected in Purpose" · Member Report</p></div>
  {'<img class="photo" src="'+member.get('photo_url','')+'" alt="Photo"/>' if member.get('photo_url') else '<div class="photo-ph">'+(member.get("name","?")[:1].upper())+'</div>'}
</div>
<div style="text-align:right;font-size:.8rem;color:#888;margin-bottom:16px;">Generated on <strong>{datetime.now().strftime('%d %b %Y, %I:%M %p')}</strong></div>
<div class="sec">Member Details</div>
<div class="ig">
  <div><label>Member ID</label><span>{member.get('id')}</span></div>
  <div><label>Full Name</label><span>{member.get('name')}</span></div>
  <div><label>Phone</label><span>{member.get('phone')}</span></div>
  <div><label>Email</label><span>{member.get('email')}</span></div>
  <div><label>University</label><span>{member.get('university')}</span></div>
  <div><label>Department</label><span>{member.get('department') or '—'}</span></div>
  <div><label>Blood Group</label><span>{member.get('blood_group') or '—'}</span></div>
  <div><label>Address</label><span>{member.get('address') or '—'}</span></div>
  <div><label>Status</label><span>{member.get('status')}</span></div>
  <div><label>Joined</label><span>{member.get('join_date')}</span></div>
</div>
<div class="sec">Payment Summary</div>
<div class="boxes">
  <div class="box" style="border-color:#27ae60;"><div class="bv" style="color:#1a5c38;">{len(ver)}</div><div class="bl">Paid</div></div>
  <div class="box" style="border-color:#f39c12;"><div class="bv" style="color:#7d6608;">{len(ped)}</div><div class="bl">Pending</div></div>
  <div class="box" style="border-color:#c0392b;"><div class="bv" style="color:#c0392b;">{len(unp)}</div><div class="bl">Due</div></div>
  <div class="box" style="border-color:#2980b9;"><div class="bv" style="color:#1a4a6b;">{tp} ৳</div><div class="bl">Total Paid</div></div>
</div>
{bal}
<div class="sec">Monthly Payment Log</div>
<table>
  <thead><tr><th>#</th><th>Month</th><th>Amount</th><th>Status</th><th>TxnID</th></tr></thead>
  <tbody>{rows}</tbody>
</table>
<div class="footer">Civision Society · {CLUB_EMAIL} · bKash/Nagad: {PAYMENT_PHONE}</div>
<div class="np">
  <button onclick="window.print()" style="padding:10px 28px;background:#1a1a1a;color:#fff;border:none;border-radius:6px;cursor:pointer;font-size:.9rem;">🖨 Print / Save PDF</button>
  <button onclick="window.close()" style="margin-left:10px;padding:10px 28px;background:#eee;border:none;border-radius:6px;cursor:pointer;font-size:.9rem;">Close</button>
</div>
</body></html>"""


@app.route("/admin/payment/update", methods=["POST"])
@admin_required
def admin_payment_update():
    mid    = request.form.get("member_id","").strip().lower()
    month  = int(request.form.get("month",0))
    year   = int(request.form.get("year",0))
    status = request.form.get("status","Verified")
    txn    = request.form.get("txn_id","").strip()
    if not mid: flash("Member ID required.","e"); return redirect(url_for("admin_dashboard")+"#tab-pay")
    pid  = f"{mid}_{month}_{year}"
    data = {"member_id":mid,"month":month,"year":year,"status":status,"txn_id":txn,
            "submitted_at":datetime.now().strftime("%Y-%m-%d %H:%M"),
            "verified_at":datetime.now().strftime("%Y-%m-%d %H:%M") if status=="Verified" else ""}
    save_payment(pid, data)
    flash(f"Payment updated for {mid}.","s")
    return redirect(url_for("admin_dashboard")+"#tab-pay")


@app.route("/admin/payment/verify", methods=["POST"])
@admin_required
def admin_payment_verify():
    pid = request.form.get("pay_id","").strip()
    update_payment(pid, {"status":"Verified","verified_at":datetime.now().strftime("%Y-%m-%d %H:%M")})
    flash("Payment verified.","s")
    return redirect(url_for("admin_dashboard")+"#tab-pay")


@app.route("/admin/payment/reject_p", methods=["POST"])
@admin_required
def admin_payment_reject():
    pid = request.form.get("pay_id","").strip()
    update_payment(pid, {"status":"Rejected"})
    flash("Payment rejected.","i")
    return redirect(url_for("admin_dashboard")+"#tab-pay")


@app.route("/admin/content/update", methods=["POST"])
@admin_required
def admin_content():
    old_content = get_home_content()
    old_notice  = old_content.get("notice","")
    new_notice  = request.form.get("notice","")
    notify_members = request.form.get("notify_members") == "1"
    col("settings").document("home").set({
        "about":         request.form.get("about",""),
        "announcements": request.form.get("announcements",""),
        "notice":        new_notice
    })
    notice_changed = new_notice.strip() != old_notice.strip() and new_notice.strip() != ""
    if notice_changed and notify_members:
        creds = get_saved_email_creds()
        se, sp = creds.get("sender_email",""), creds.get("sender_password","")
        if se and sp:
            members = [m for m in get_all_members() if m.get("status")=="Approved" and m.get("email")]
            sent = failed = 0
            try:
                with smtplib.SMTP("smtp.gmail.com",587) as srv:
                    srv.ehlo(); srv.starttls(); srv.login(se,sp)
                    for m in members:
                        try:
                            subj = "New Notice — Civision Society"
                            plain = (f"Dear {m.get('name')},\n\nA new notice has been posted:\n\n"
                                     f"{new_notice}\n\nCivision Society\n{CLUB_EMAIL}")
                            html = (f'<div style="font-family:sans-serif;max-width:560px;margin:auto;padding:24px;">'
                                    f'<h2 style="font-family:Georgia,serif;color:#1a1a1a;border-bottom:2px solid #C9A84C;padding-bottom:8px;">📢 New Notice</h2>'
                                    f'<p>Dear <strong>{m.get("name")}</strong>,</p>'
                                    f'<div style="background:#fef9e7;border:1px solid #f9e79f;border-radius:8px;padding:18px;margin:16px 0;">'
                                    f'<p style="margin:0;line-height:1.7;">{new_notice}</p></div>'
                                    f'<p style="color:#aaa;font-size:.78rem;text-align:center;">{CLUB_EMAIL} | ID: {m.get("id")}</p></div>')
                            msg = MIMEMultipart("alternative")
                            msg["Subject"]=subj; msg["From"]=se; msg["To"]=m["email"]
                            msg.attach(MIMEText(plain,"plain")); msg.attach(MIMEText(html,"html"))
                            srv.sendmail(se, m["email"], msg.as_string()); sent+=1
                        except: failed+=1
                flash(f"Content updated. Notice emailed: {sent} sent, {failed} failed.","s")
            except Exception as e:
                flash(f"Content updated, but notice email failed: {str(e)}","e")
        else:
            flash("Content updated. Notice NOT emailed — save Gmail credentials in the Email tab first.","i")
    else:
        flash("Content updated.","s")
    return redirect(url_for("admin_dashboard")+"#tab-cnt")


@app.route("/admin/gallery/upload", methods=["POST"])
@admin_required
def admin_gallery_upload():
    if "image" not in request.files: flash("No file.","e"); return redirect(url_for("admin_dashboard")+"#tab-gal")
    f = request.files["image"]
    caption = request.form.get("caption","").strip()
    if f and allowed_file(f.filename):
        url = upload_image(f, folder="civision/gallery")
        gid = f"gal_{int(datetime.now().timestamp())}"
        col("gallery").document(gid).set({"url":url,"caption":caption,"uploaded_at":datetime.now().strftime("%Y-%m-%d %H:%M")})
        flash("Image uploaded.","s")
    else: flash("Invalid file.","e")
    return redirect(url_for("admin_dashboard")+"#tab-gal")


@app.route("/admin/gallery/delete", methods=["POST"])
@admin_required
def admin_gallery_delete():
    img_id = request.form.get("image_id","").strip()
    col("gallery").document(img_id).delete()
    flash("Image deleted.","i")
    return redirect(url_for("admin_dashboard")+"#tab-gal")


@app.route("/admin/events/add", methods=["POST"])
@admin_required
def admin_events_add():
    title   = request.form.get("title","").strip()
    desc    = request.form.get("description","").strip()
    ev_date = request.form.get("event_date","").strip()
    img_url = ""
    if "image" in request.files:
        f = request.files["image"]
        if f and f.filename and allowed_file(f.filename):
            img_url = upload_image(f, folder="civision/events")
    eid = f"ev_{int(datetime.now().timestamp())}"
    col("events").document(eid).set({
        "title":title,"description":desc,"event_date":ev_date,
        "image_url":img_url,"created_at":datetime.now().strftime("%Y-%m-%d")
    })
    flash(f"Event '{title}' added.","s")
    return redirect(url_for("admin_dashboard")+"#tab-ev")


@app.route("/admin/events/delete", methods=["POST"])
@admin_required
def admin_events_delete():
    eid = request.form.get("event_id","").strip()
    col("events").document(eid).delete()
    flash("Event deleted.","i")
    return redirect(url_for("admin_dashboard")+"#tab-ev")


@app.route("/admin/expenses/add", methods=["POST"])
@admin_required
def admin_expenses_add():
    title  = request.form.get("title","").strip()
    amount = request.form.get("amount","0")
    dt     = request.form.get("date","")
    notes  = request.form.get("notes","").strip()
    eid    = f"exp_{int(datetime.now().timestamp())}"
    col("expenses").document(eid).set({"title":title,"amount":float(amount),"date":dt,"notes":notes})
    flash("Expense added.","s")
    return redirect(url_for("admin_dashboard")+"#tab-sum")


@app.route("/admin/expenses/delete", methods=["POST"])
@admin_required
def admin_expenses_delete():
    eid = request.form.get("expense_id","").strip()
    col("expenses").document(eid).delete()
    flash("Expense deleted.","i")
    return redirect(url_for("admin_dashboard")+"#tab-sum")


@app.route("/admin/logo/upload", methods=["POST"])
@admin_required
def admin_logo_upload():
    if "logo" not in request.files: flash("No file.","e"); return redirect(url_for("admin_dashboard")+"#tab-logo")
    f = request.files["logo"]
    if not f or not allowed_file(f.filename): flash("Invalid file.","e"); return redirect(url_for("admin_dashboard")+"#tab-logo")
    url = upload_image(f, folder="civision/logo")
    col("settings").document("logo").set({"url":url})
    flash("Logo updated! It will also appear as the site favicon.","s")
    return redirect(url_for("admin_dashboard")+"#tab-logo")


@app.route("/admin/settings/bg_upload", methods=["POST"])
@admin_required
def admin_bg_upload():
    if "bg_image" not in request.files: flash("No file.","e"); return redirect(url_for("admin_dashboard")+"#tab-logo")
    f = request.files["bg_image"]
    if not f or not allowed_file(f.filename): flash("Invalid file.","e"); return redirect(url_for("admin_dashboard")+"#tab-logo")
    url = upload_image(f, folder="civision/background")
    site = get_site_settings()
    site["bg_image"] = url
    col("settings").document("site").set(site)
    flash("Background image updated!","s")
    return redirect(url_for("admin_dashboard")+"#tab-logo")


@app.route("/admin/settings/bg_remove", methods=["POST"])
@admin_required
def admin_bg_remove():
    site = get_site_settings()
    site["bg_image"] = ""
    col("settings").document("site").set(site)
    flash("Background image removed.","i")
    return redirect(url_for("admin_dashboard")+"#tab-logo")


@app.route("/admin/settings/mobile_scale", methods=["POST"])
@admin_required
def admin_mobile_scale():
    scale = request.form.get("mobile_scale","1.0").strip()
    try:
        val = float(scale)
        if val < 0.5 or val > 2.0: raise ValueError()
    except ValueError:
        flash("Scale must be a number between 0.5 and 2.0.","e")
        return redirect(url_for("admin_dashboard")+"#tab-logo")
    site = get_site_settings()
    site["mobile_scale"] = str(val)
    col("settings").document("site").set(site)
    flash("Mobile scale updated!","s")
    return redirect(url_for("admin_dashboard")+"#tab-logo")


@app.route("/admin/change_password", methods=["POST"])
@admin_required
def admin_change_password():
    cur  = request.form.get("current_password","").strip()
    new  = request.form.get("new_password","").strip()
    conf = request.form.get("confirm_password","").strip()
    if not check_admin_password(cur):
        flash("Current password incorrect.","e")
        return redirect(url_for("admin_dashboard")+"#tab-logo")
    if len(new) < 4:
        flash("New password too short (min 4 chars).","e")
        return redirect(url_for("admin_dashboard")+"#tab-logo")
    if new != conf:
        flash("Passwords do not match.","e")
        return redirect(url_for("admin_dashboard")+"#tab-logo")
    set_admin_pw_hash(hashlib.sha256(new.encode()).hexdigest())
    flash("Admin password changed successfully!","s")
    return redirect(url_for("admin_dashboard")+"#tab-logo")


# ── DB EDITOR API ──────────────────────────────────────────────────────────────
@app.route("/admin/db/member/<mid>")
@admin_required
def admin_db_get_member(mid):
    member = get_member_by_id(mid.lower())
    if not member: return jsonify({"error": f"Member '{mid}' not found."})
    safe = {k:v for k,v in member.items() if k != "password_hash"}
    return jsonify({"member": safe})


@app.route("/admin/db/member/save", methods=["POST"])
@admin_required
def admin_db_save_member():
    original_id = request.form.get("original_id","").strip().lower()
    new_id      = request.form.get("id","").strip().lower()
    new_pw      = request.form.get("new_password","").strip()
    fields = ["name","phone","email","university","department","blood_group","dob","address","status","join_date"]
    data   = {f: request.form.get(f,"").strip() for f in fields}
    data["id"] = new_id
    existing = get_member_by_id(original_id)
    if new_pw:
        data["password_hash"]  = hash_pw(new_pw)
        data["session_token"]  = ""   # force re-login on the device after a password reset
    else:
        if existing: data["password_hash"] = existing.get("password_hash","")
    if existing:
        data["photo_url"]     = existing.get("photo_url","")
        data.setdefault("session_token", existing.get("session_token",""))
    if new_id != original_id:
        col("members").document(original_id).delete()
    save_member(new_id, data)
    flash(f"Member {new_id} saved successfully.","s")
    return redirect(url_for("admin_dashboard")+"#tab-db")


@app.route("/admin/db/payments/<mid>")
@admin_required
def admin_db_get_payments(mid):
    payments = get_payments_for_member(mid.lower())
    return jsonify({"payments": payments})


@app.route("/admin/db/payment/update", methods=["POST"])
@admin_required
def admin_db_payment_update():
    pid    = request.form.get("pay_id","").strip()
    status = request.form.get("status","Pending")
    txn    = request.form.get("txn_id","").strip()
    update_payment(pid, {
        "status": status, "txn_id": txn,
        "verified_at": datetime.now().strftime("%Y-%m-%d %H:%M") if status=="Verified" else ""
    })
    flash("Payment record updated.","s")
    return redirect(url_for("admin_dashboard")+"#tab-db")


@app.route("/admin/db/payment/delete", methods=["POST"])
@admin_required
def admin_db_payment_delete():
    pid = request.form.get("pay_id","").strip()
    delete_payment(pid)
    flash("Payment record deleted.","i")
    return redirect(url_for("admin_dashboard")+"#tab-db")


# ── DUE NOTICES ────────────────────────────────────────────────────────────────
def _build_formal_due_email(member, unpaid):
    mid   = member.get("id","")
    name  = member.get("name","")
    email = member.get("email","")
    uni   = member.get("university","") or "—"
    dept  = member.get("department","") or "—"
    today = datetime.now().strftime("%d %B %Y")
    total_due = len([p for p in unpaid if p.get("status","")=="Unpaid"]) * MONTHLY_DUES
    subj  = f"Formal Payment Due Notice — Civision Society"
    plain = (f"CIVISION SOCIETY — Connected in Purpose\n"
             f"{CLUB_EMAIL} | bKash/Nagad: {PAYMENT_PHONE}\n"
             f"{'='*60}\n"
             f"Date: {today}\n\n"
             f"Dear {name},\n\n"
             f"This is a formal reminder from the Administration of Civision Society\n"
             f"regarding your outstanding monthly membership dues.\n\n"
             f"Outstanding Months:\n"
             + "".join(f"  • {p.get('month','')} — ৳{MONTHLY_DUES} [{p.get('status','')}]\n" for p in unpaid)
             + f"\nTotal Outstanding: ৳{total_due}\n\n"
             f"PAYMENT INSTRUCTIONS:\n"
             f"  Send ৳{MONTHLY_DUES}/month via bKash or Nagad to: {PAYMENT_PHONE}\n"
             f"  Then submit your Transaction ID on the Payment page.\n\n"
             f"Failure to clear dues may result in suspension of membership.\n\n"
             f"Yours sincerely,\nAdministration — Civision Society\n{CLUB_EMAIL}")
    rows_html = "".join(
        f'<tr><td style="padding:9px 14px;">{i+1}</td>'
        f'<td style="padding:9px 14px;">{p.get("month","")}</td>'
        f'<td style="padding:9px 14px;">৳{MONTHLY_DUES}</td>'
        f'<td style="padding:9px 14px;color:{"#c0392b" if p.get("status")=="Unpaid" else "#856404"};font-weight:600;">'
        f'{"Unpaid" if p.get("status")=="Unpaid" else "Pending Verification"}</td></tr>'
        for i,p in enumerate(unpaid))
    html = f"""<!DOCTYPE html><html><head><meta charset="UTF-8"/>
<style>
body{{font-family:Georgia,serif;max-width:680px;margin:0 auto;padding:0;color:#1a1a1a;background:#fff;}}
.lh{{background:linear-gradient(135deg,#1a1a1a,#2c3e50);color:#fff;padding:28px 36px;}}
.lh h1{{font-size:1.8rem;font-weight:700;letter-spacing:.04em;margin:0 0 4px;}}
.lh p{{color:rgba(255,255,255,.65);font-size:.82rem;margin:0;}}
.lb{{padding:32px 36px;}}
.dr{{display:flex;justify-content:space-between;margin-bottom:20px;font-size:.83rem;color:#666;}}
.to{{background:#f9f9f7;border-left:4px solid #C9A84C;padding:14px 18px;margin-bottom:20px;border-radius:0 6px 6px 0;line-height:1.7;font-size:.9rem;}}
.subj{{font-size:1rem;font-weight:700;border-bottom:2px solid #C9A84C;padding-bottom:8px;margin-bottom:18px;}}
.bt{{font-size:.92rem;line-height:1.9;color:#2a2a2a;}}.bt p{{margin-bottom:12px;}}
table.dt{{width:100%;border-collapse:collapse;margin:18px 0;}}
table.dt th{{background:#1a1a1a;color:#fff;padding:10px 14px;text-align:left;font-size:.8rem;text-transform:uppercase;}}
table.dt td{{border-bottom:1px solid #eee;}}
table.dt tr.tot td{{font-weight:700;background:#fef9e7;border-bottom:none;padding:10px 14px;}}
.pb{{background:#EBF5FB;border:1px solid #AED6F1;border-radius:8px;padding:18px 20px;margin:18px 0;}}
.pb h3{{font-size:.9rem;font-weight:700;margin:0 0 8px;color:#1a3a55;}}
.pb .num{{font-size:1.4rem;font-weight:700;letter-spacing:.06em;}}
.warn{{background:#fef9e7;border:1px solid #f9e79f;border-radius:6px;padding:12px 16px;margin:14px 0;font-size:.85rem;color:#7d6608;}}
.sig{{margin-top:32px;padding-top:18px;border-top:1px solid #ddd;font-size:.9rem;line-height:1.8;}}
.fb{{background:#1a1a1a;color:rgba(255,255,255,.5);text-align:center;padding:12px;font-size:.75rem;}}
.fb span{{color:#C9A84C;}}
</style></head><body>
<div class="lh">
  <h1>Civision Society</h1>
  <p>"Connected in Purpose" &nbsp;·&nbsp; {CLUB_EMAIL} &nbsp;·&nbsp; bKash/Nagad: {PAYMENT_PHONE}</p>
</div>
<div class="lb">
  <div class="dr">
    <span>Date: <strong>{today}</strong></span>
  </div>
  <div class="bt">
    <p>Dear <strong>{name}</strong>,</p>
    <p>This is a <strong>formal reminder</strong> from the Administration of <strong>Civision Society</strong> regarding your outstanding monthly membership dues. As per our official records, the following months remain unpaid or pending verification:</p>
    <table class="dt">
      <thead><tr><th>#</th><th>Month</th><th>Amount</th><th>Status</th></tr></thead>
      <tbody>
        {rows_html}
        <tr class="tot"><td colspan="2">Total Outstanding</td><td colspan="2">৳{total_due}</td></tr>
      </tbody>
    </table>
    <p>You are kindly requested to clear the above dues at your earliest convenience to maintain your active membership status.</p>
    <div class="pb">
      <h3>💳 Payment Instructions</h3>
      <p>Send <strong>৳{MONTHLY_DUES} per month</strong> via bKash or Nagad to:</p>
      <div class="num">{PAYMENT_PHONE}</div>
      <p style="margin-top:8px;font-size:.85rem;color:#555;">After payment, submit your <strong>Transaction ID</strong> on the Payment page of our website.</p>
    </div>
    <div class="warn">⚠️ Failure to clear outstanding dues may result in suspension of membership privileges until the balance is settled.</div>
    <p>If you have already made the payment, please submit your Transaction ID immediately for verification. For queries: <a href="mailto:{CLUB_EMAIL}">{CLUB_EMAIL}</a></p>
  </div>
  <div class="sig">
    <p>Yours sincerely,</p>
    <p style="margin-top:12px;"><strong>Administration</strong><br/><strong>Civision Society</strong><br/>
    <em>"Connected in Purpose"</em><br/>{CLUB_EMAIL} &nbsp;|&nbsp; bKash/Nagad: {PAYMENT_PHONE}<br/>Date: {today}</p>
  </div>
</div>
<div class="fb"><span>Civision Society</span> — Official Correspondence — {today}</div>
</body></html>"""
    return subj, plain, html


@app.route("/admin/notices/send_one", methods=["POST"])
@admin_required
def admin_notice_one():
    mid = request.form.get("member_id","").strip().lower()
    creds = get_saved_email_creds()
    se, sp = creds.get("sender_email",""), creds.get("sender_password","")
    if not se or not sp:
        flash("Save your Gmail credentials in the Email tab first.","e")
        return redirect(url_for("admin_dashboard")+"#tab-sum")
    member = get_member_by_id(mid)
    if not member or not member.get("email"):
        flash("Member or email not found.","e")
        return redirect(url_for("admin_dashboard")+"#tab-sum")
    ps     = get_payment_status_for_member(mid)
    unpaid = [p for p in ps if p["status"] in ("Unpaid","Pending")]
    if not unpaid:
        flash(f"{member.get('name')} has no outstanding dues.","i")
        return redirect(url_for("admin_dashboard")+"#tab-sum")
    subj, plain, html = _build_formal_due_email(member, unpaid)
    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subj; msg["From"] = se; msg["To"] = member["email"]
        msg.attach(MIMEText(plain,"plain")); msg.attach(MIMEText(html,"html"))
        with smtplib.SMTP("smtp.gmail.com",587) as srv:
            srv.ehlo(); srv.starttls(); srv.login(se,sp)
            srv.sendmail(se, member["email"], msg.as_string())
        flash(f"✓ Formal notice sent to {member.get('name')} ({member['email']}).","s")
    except Exception as e:
        flash(f"Failed to send: {str(e)}","e")
    return redirect(url_for("admin_dashboard")+"#tab-sum")


@app.route("/admin/notices/send_all", methods=["POST"])
@admin_required
def admin_notice_all():
    creds = get_saved_email_creds()
    se, sp = creds.get("sender_email",""), creds.get("sender_password","")
    if not se or not sp:
        flash("Save your Gmail credentials in the Email tab first.","e")
        return redirect(url_for("admin_dashboard")+"#tab-sum")
    members = [m for m in get_all_members() if m.get("status")=="Approved" and m.get("email")]
    sent=failed=skipped=0
    try:
        with smtplib.SMTP("smtp.gmail.com",587) as srv:
            srv.ehlo(); srv.starttls(); srv.login(se,sp)
            for m in members:
                ps     = get_payment_status_for_member(m["id"])
                unpaid = [p for p in ps if p["status"] in ("Unpaid","Pending")]
                if not unpaid: skipped+=1; continue
                subj, plain, html = _build_formal_due_email(m, unpaid)
                try:
                    msg = MIMEMultipart("alternative")
                    msg["Subject"]=subj; msg["From"]=se; msg["To"]=m["email"]
                    msg.attach(MIMEText(plain,"plain")); msg.attach(MIMEText(html,"html"))
                    srv.sendmail(se, m["email"], msg.as_string()); sent+=1
                except: failed+=1
    except Exception as e:
        flash(f"SMTP failed: {str(e)}","e")
        return redirect(url_for("admin_dashboard")+"#tab-sum")
    flash(f"Done: {sent} sent, {failed} failed, {skipped} already up-to-date.","s")
    return redirect(url_for("admin_dashboard")+"#tab-sum")


@app.route("/admin/email/save_creds", methods=["POST"])
@admin_required
def admin_save_email_creds():
    se = request.form.get("sender_email","").strip()
    sp = request.form.get("sender_password","").strip()
    if not se or not sp:
        flash("Both Gmail address and App Password are required.","e")
        return redirect(url_for("admin_dashboard")+"#tab-eml")
    save_email_creds(se, sp)
    flash("✓ Gmail credentials saved and will be remembered.","s")
    return redirect(url_for("admin_dashboard")+"#tab-eml")


@app.route("/admin/email/save_notify", methods=["POST"])
@admin_required
def admin_save_notify_email():
    email = request.form.get("notify_email","").strip()
    if not email:
        flash("Enter an email address.","e")
        return redirect(url_for("admin_dashboard")+"#tab-eml")
    save_notify_email(email)
    flash("✓ Notification email saved.","s")
    return redirect(url_for("admin_dashboard")+"#tab-eml")


@app.route("/admin/email/broadcast", methods=["POST"])
@admin_required
def admin_broadcast():
    creds = get_saved_email_creds()
    se, sp = creds.get("sender_email",""), creds.get("sender_password","")
    subj = request.form.get("subject","").strip()
    body = request.form.get("body","").strip()
    if not se or not sp:
        flash("Save your Gmail credentials above first.","e")
        return redirect(url_for("admin_dashboard")+"#tab-eml")
    if not all([subj,body]):
        flash("Subject and message are required.","e")
        return redirect(url_for("admin_dashboard")+"#tab-eml")
    members = [m for m in get_all_members() if m.get("status")=="Approved" and m.get("email")]
    if not members:
        flash("No approved members with email found.","e")
        return redirect(url_for("admin_dashboard")+"#tab-eml")
    sent=failed=0
    try:
        with smtplib.SMTP("smtp.gmail.com",587) as srv:
            srv.ehlo(); srv.starttls(); srv.login(se,sp)
            for m in members:
                try:
                    html = (f'<div style="font-family:sans-serif;max-width:560px;margin:auto;padding:24px;">'
                            f'<h2 style="font-family:Georgia,serif;color:#1a1a1a;">Civision Society</h2>'
                            f'<p>Dear <strong>{m.get("name")}</strong>,</p>'
                            f'<div style="background:#f9f9f7;border-radius:8px;padding:20px;border-left:4px solid #C9A84C;margin:16px 0;">'
                            f'<div style="line-height:1.7;">{body.replace(chr(10),"<br/>")}</div></div>'
                            f'<p style="color:#aaa;font-size:.78rem;text-align:center;">{CLUB_EMAIL} | ID: {m.get("id")}</p></div>')
                    msg = MIMEMultipart("alternative")
                    msg["Subject"]=subj; msg["From"]=se; msg["To"]=m["email"]
                    msg.attach(MIMEText(body,"plain")); msg.attach(MIMEText(html,"html"))
                    srv.sendmail(se, m["email"], msg.as_string()); sent+=1
                except: failed+=1
    except Exception as e:
        flash(f"SMTP failed: {str(e)}","e")
        return redirect(url_for("admin_dashboard")+"#tab-eml")
    flash(f"Broadcast done: {sent} sent, {failed} failed.","s")
    return redirect(url_for("admin_dashboard")+"#tab-eml")


@app.route("/googled600bf5cd02a9099.html")
def google_site_verification():
    return "google-site-verification: googled600bf5cd02a9099.html", 200, {"Content-Type": "text/html"}


@app.route("/googled600bf5cd02a9099.html")
def google_verify():
    return "google-site-verification: googled600bf5cd02a9099.html", 200, {"Content-Type": "text/html"}


@app.route("/robots.txt")
def robots_txt():
    base = request.url_root.rstrip("/")
    body = (f"User-agent: *\n"
            f"Allow: /\n"
            f"Allow: /join\n"
            f"Allow: /login\n"
            f"Allow: /payment\n"
            f"Allow: /admin\n"
            f"Disallow: /admin/dashboard\n"
            f"Disallow: /member/profile\n"
            f"Sitemap: {base}/sitemap.xml\n")
    return body, 200, {"Content-Type": "text/plain"}


@app.route("/sitemap.xml")
def sitemap_xml():
    base = request.url_root.rstrip("/")
    pages = [
        ("/", "1.0", "daily"),
        ("/login", "0.8", "monthly"),
        ("/join", "0.8", "monthly"),
        ("/payment", "0.8", "monthly"),
        ("/admin", "0.5", "monthly"),
    ]
    urls = "".join(
        f"<url><loc>{base}{p}</loc><changefreq>{freq}</changefreq><priority>{pri}</priority></url>"
        for p, pri, freq in pages
    )
    xml = f'<?xml version="1.0" encoding="UTF-8"?><urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">{urls}</urlset>'
    return xml, 200, {"Content-Type": "application/xml"}


@app.route("/favicon.ico")
def favicon_ico():
    logo = get_nav_logo()
    if logo:
        return redirect(logo)
    return "", 404


if __name__ == "__main__":
    app.run(debug=False, port=5000)
